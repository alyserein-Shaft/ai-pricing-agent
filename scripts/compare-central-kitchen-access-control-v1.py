from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
FINAL_QUOTATION = BASE / "final-quotation.pdf"

OUTPUT_DIR = Path("outputs/central-kitchen-access-control")
OUTPUT_FILE = OUTPUT_DIR / "Central_Kitchen_Access_Control_Comparison_v1.xlsx"


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()

    replacements = {
        "card access": "access control",
        "card reader": "reader",
        "door reader": "reader",
        "pushbutton": "push button",
        "magnetic contact": "door contact",
        "magnetic door contact": "door contact",
        "access-control": "access control",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)

    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def longest_text(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if isinstance(value, str) and len(clean(value)) >= 5
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    numbers = []

    for value in values:
        number = numeric(value)

        if number is not None:
            numbers.append(number)

    if not numbers:
        return None

    for number in reversed(numbers):
        if number >= 1:
            return number

    return numbers[-1]


def find_unit(values: list[Any]) -> str:
    known_units = {
        "no",
        "nos",
        "no.",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in known_units:
            return clean(value)

    return ""


def find_item_number(values: list[Any]) -> str:
    for value in values[:6]:
        text = clean(value)

        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
            return text

    return ""


def classify_access_item(description: str) -> str:
    text = normalize(description)

    if "controller" in text or "control panel" in text:
        return "Access Controller"

    if "reader" in text:
        return "Card Reader"

    if "push button" in text or "request to exit" in text or "rex" in text:
        return "Push Button / REX"

    if "door contact" in text or "magnetic contact" in text:
        return "Door Contact"

    if "magnetic lock" in text or "maglock" in text:
        return "Magnetic Lock"

    if "electric strike" in text:
        return "Electric Strike"

    if "break glass" in text:
        return "Emergency Break Glass"

    if "power supply" in text or "psu" in text:
        return "Power Supply"

    if "access control" in text and "software" in text:
        return "Access Control Software"

    if "license" in text:
        return "Access Control License"

    if "testing" in text or "commissioning" in text:
        return "Testing & Commissioning"

    return ""


def extract_source_access_control() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if "28" not in sheet_name and "access" not in sheet_name.lower():
            continue

        ws = workbook[sheet_name]
        inside_access = False

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))
            normalized = normalize(joined)

            if not joined:
                continue

            if any(
                phrase in normalized
                for phrase in (
                    "access control system",
                    "card access system",
                )
            ):
                inside_access = True

            if inside_access and any(
                phrase in normalized
                for phrase in (
                    "fire alarm system",
                    "public address system",
                    "closed circuit television",
                    "cctv system",
                    "data networking",
                )
            ):
                if "access control" not in normalized:
                    break

            if not inside_access:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)

            if not description or quantity is None:
                continue

            category = classify_access_item(description)

            if not category:
                continue

            records.append(
                {
                    "source_category": category,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "item_number": find_item_number(values),
                    "description": description,
                    "quantity": quantity,
                    "unit": find_unit(values),
                    "raw_row": joined,
                }
            )

    return records


def extract_pdf_lines() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    lines: list[dict[str, Any]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            line = clean(raw_line)

            if line:
                lines.append(
                    {
                        "page": page_number,
                        "line_number": line_number,
                        "text": line,
                    }
                )

    return lines


def parse_amounts(text: str) -> tuple[float | None, float | None, float | None]:
    matches = list(
        re.finditer(
            r"(?<!\d)"
            r"(\d+(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
            r"(?:\s|$)",
            text,
        )
    )

    if not matches:
        return None, None, None

    match = matches[-1]

    return (
        float(match.group(1)),
        float(match.group(2).replace(",", "")),
        float(match.group(3).replace(",", "")),
    )


def extract_part_number(description: str) -> str:
    text = clean(description)

    candidates = re.findall(
        r"\b[A-Z0-9][A-Z0-9()_.]*"
        r"(?:[-/][A-Z0-9()_.]+)+\b",
        text[:220],
        flags=re.IGNORECASE,
    )

    rejected = (
        r"^\d+VDC",
        r"^\d+VAC",
        r"^\d+A$",
        r"^Q1067",
    )

    for candidate in candidates:
        if any(
            re.match(pattern, candidate, flags=re.IGNORECASE)
            for pattern in rejected
        ):
            continue

        if re.search(r"[A-Za-z]", candidate) and re.search(r"\d", candidate):
            return candidate

    return ""


def is_item_start(text: str) -> bool:
    match = re.match(r"^\s*(\d{1,2})\s+(.+)$", text)

    if not match:
        return False

    serial = int(match.group(1))
    remainder = match.group(2)

    if not 1 <= serial <= 50:
        return False

    return bool(re.search(r"[A-Za-z]", remainder))


def extract_final_access_control() -> list[dict[str, Any]]:
    lines = extract_pdf_lines()
    section_lines = []
    inside_access = False

    start_markers = (
        "access control system",
        "card access system",
    )

    stop_markers = (
        "fire alarm system",
        "public address system",
        "closed circuit television",
        "cctv system",
        "data networking",
        "structured cabling",
    )

    for entry in lines:
        normalized = normalize(entry["text"])

        if any(marker in normalized for marker in start_markers):
            inside_access = True
            continue

        if inside_access and any(marker in normalized for marker in stop_markers):
            inside_access = False
            break

        if inside_access:
            section_lines.append(entry)

    blocks = []
    current = None

    for entry in section_lines:
        text = entry["text"]

        if is_item_start(text):
            if current:
                blocks.append(current)

            match = re.match(r"^\s*(\d{1,2})\s+(.+)$", text)

            current = {
                "serial_number": int(match.group(1)),
                "page": entry["page"],
                "lines": [match.group(2)],
            }

        elif current:
            current["lines"].append(text)

    if current:
        blocks.append(current)

    records = []

    for block in blocks:
        description = clean(" ".join(block["lines"]))
        quantity, unit_price, item_price = parse_amounts(description)
        part_number = extract_part_number(description)
        category = classify_access_item(f"{part_number} {description}")

        if not category:
            category = "Other Access Control Component"

        records.append(
            {
                "serial_number": block["serial_number"],
                "page": block["page"],
                "final_category": category,
                "part_number": part_number,
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
                "item_price": item_price,
            }
        )

    return records


def build_relationships(
    source_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    final_by_category: dict[str, list[dict[str, Any]]] = {}

    for record in final_records:
        final_by_category.setdefault(
            record["final_category"],
            [],
        ).append(record)

    relationships = []

    for source in source_records:
        category = source["source_category"]
        matches = final_by_category.get(category, [])

        relationships.append(
            {
                "source_category": category,
                "source_description": source["description"],
                "source_quantity": source["quantity"],
                "source_unit": source["unit"],
                "final_categories": "\n".join(
                    record["final_category"] for record in matches
                ),
                "final_part_numbers": "\n".join(
                    record["part_number"]
                    for record in matches
                    if record["part_number"]
                ),
                "final_quantities": ", ".join(
                    str(record["quantity"])
                    for record in matches
                    if record["quantity"] is not None
                ),
                "relationship_type": (
                    "Direct Mapping"
                    if len(matches) == 1
                    else "One-to-Many / Needs Review"
                ),
                "engineering_observation": "",
                "candidate_learning_rule": "",
            }
        )

    return relationships


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = max(
            len(clean(cell.value))
            for cell in column_cells
        )

        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(
            max(max_length + 2, 12),
            55,
        )


def main() -> None:
    for path in (SOURCE_BOQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_access_control()
    final_records = extract_final_access_control()
    relationships = build_relationships(source_records, final_records)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Source Access Control"),
        [
            "source_category",
            "sheet",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        source_records,
    )

    write_sheet(
        workbook.create_sheet("Final Access Control"),
        [
            "serial_number",
            "page",
            "final_category",
            "part_number",
            "description",
            "quantity",
            "unit_price",
            "item_price",
        ],
        final_records,
    )

    write_sheet(
        workbook.create_sheet("Relationships"),
        [
            "source_category",
            "source_description",
            "source_quantity",
            "source_unit",
            "final_categories",
            "final_part_numbers",
            "final_quantities",
            "relationship_type",
            "engineering_observation",
            "candidate_learning_rule",
        ],
        relationships,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source Access Control items: {len(source_records)}")
    print(f"Final Access Control line items: {len(final_records)}")
    print(f"Relationship rows: {len(relationships)}")
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
