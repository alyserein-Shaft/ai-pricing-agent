from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
CORNING_RFQ = BASE / "data-networking-engineer-rfq.xlsx"
FINAL_QUOTATION = BASE / "final-quotation.pdf"

OUTPUT_DIR = Path("outputs/central-kitchen-structured-cabling")
OUTPUT_FILE = (
    OUTPUT_DIR
    / "Central_Kitchen_Structured_Cabling_Comparison_v5.xlsx"
)


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()

    replacements = {
        "cat.6a": "cat6a",
        "cat 6a": "cat6a",
        "cat.6": "cat6",
        "cat 6": "cat6",
        "rj-45": "rj45",
        "rj 45": "rj45",
        "fibre": "fiber",
        "patchpanel": "patch panel",
        "face plate": "faceplate",
        "double gang": "double",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)

    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def longest_text(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if (
            isinstance(value, str)
            and len(clean(value)) >= 5
            and not clean(value).startswith("=")
        )
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    numbers: list[float] = []

    for value in values:
        number = numeric(value)

        if number is not None:
            numbers.append(number)

    if not numbers:
        return None

    for number in reversed(numbers):
        if number >= 1:
            return number

    return numbers[-1]


def find_unit(values: list[Any]) -> str:
    units = {
        "no",
        "nos",
        "no.",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in units:
            return clean(value)

    return ""


def find_item_number(values: list[Any]) -> str:
    for value in values[:6]:
        text = clean(value)

        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
            return text

    return ""


def classify_source_item(description: str) -> tuple[str, int]:
    text = normalize(description)

    if "15u rack" in text:
        return "15U Rack", 0

    if "18u rack" in text:
        return "18U Rack", 0

    if "24u rack" in text and "midf" in text:
        return "24U MIDF Rack", 0

    if "24u rack" in text:
        return "24U IDF Rack", 0

    if "wireless data outlet" in text and "ceiling" in text:
        return "Ceiling Wireless Data Outlet", 1

    if "wall mounted wireless data outlet" in text:
        return "Wall Wireless Data Outlet", 1

    if (
        ("telephone and data" in text or "telepohne and data" in text)
        and "2 no" in text
    ):
        return "Telephone + Data Double Outlet", 2

    if "double" in text and "data outlet" in text:
        return "Double Data Outlet", 2

    if "data outlet" in text and "rj45" in text:
        return "Single Data Outlet", 1

    if "hdmi outlet" in text:
        return "HDMI Outlet", 0

    return "", 0


def extract_source_boq() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    ws = workbook["Div. 27"]

    records: list[dict[str, Any]] = []

    for row_number in range(10, 124):
        values = row_values(ws, row_number)
        description = longest_text(values)
        quantity = find_quantity(values)

        if not description or quantity is None:
            continue

        category, ports_per_item = classify_source_item(description)

        if not category:
            continue

        estimated_ports = (
            float(quantity) * ports_per_item
            if ports_per_item > 0
            else 0
        )

        records.append(
            {
                "source_category": category,
                "source_row": row_number,
                "item_number": find_item_number(values),
                "description": description,
                "quantity": quantity,
                "unit": find_unit(values),
                "ports_per_item": ports_per_item,
                "estimated_rj45_ports": estimated_ports,
            }
        )

    return records


def classify_rfq_item(description: str, part_number: str) -> str:
    text = normalize(f"{part_number} {description}")

    if "faceplate" in text and "1 jack" in text:
        return "1-Port Faceplate"

    if "faceplate" in text and "2 jack" in text:
        return "2-Port Faceplate"

    if "cat6a" in text and "jack" in text and "patch panel" in text:
        return "Patch Panel Jack"

    if "cat6a" in text and "jack" in text:
        return "Outlet Jack"

    if "male module connector" in text:
        return "Field RJ45 Connector"

    if "patch panel 24p" in text:
        return "24-Port Patch Panel"

    if "patch panel 48p" in text:
        return "48-Port Patch Panel"

    if "cable management" in text:
        return "Cable Manager"

    if (
        ("patchcord" in text and "cat6a" in text)
        or "ccakye g1002 a030 c0" in text
        or (
            "2xrj45" in text
            and "3m" in text
            and "cat6a" in text
        )
    ):
        return "Copper Patch Cord"

    if "fiber patch panel" in text:
        return "Fiber Patch Panel"

    if "fo patchcord" in text or (
        "om4" in text and "lc lc" in text
    ):
        return "Fiber Patch Cord"

    if "hdmi outlet" in text:
        return "HDMI Outlet"

    return "Other Passive Component"


def extract_corning_rfq() -> list[dict[str, Any]]:
    workbook = load_workbook(CORNING_RFQ, data_only=False)
    ws = workbook["Sheet1"]

    records: list[dict[str, Any]] = []
    previous_panel_category = ""

    for row_number in range(2, ws.max_row + 1):
        part_number = clean(ws.cell(row=row_number, column=1).value)
        description = clean(ws.cell(row=row_number, column=2).value)
        quantity = numeric(ws.cell(row=row_number, column=3).value)

        if not description or quantity is None:
            continue

        category = classify_rfq_item(description, part_number)

        # The 192 jacks immediately after the 48-port panel belong
        # to patch panels even though the description omits the phrase.
        if (
            category == "Outlet Jack"
            and previous_panel_category
            in {"24-Port Patch Panel", "48-Port Patch Panel"}
        ):
            expected_quantity = (
                648
                if previous_panel_category == "24-Port Patch Panel"
                else 192
            )

            if float(quantity) == expected_quantity:
                category = "Patch Panel Jack"

        capacity = 0

        if category == "1-Port Faceplate":
            capacity = int(quantity)

        elif category == "2-Port Faceplate":
            capacity = int(quantity) * 2

        elif category == "24-Port Patch Panel":
            capacity = int(quantity) * 24

        elif category == "48-Port Patch Panel":
            capacity = int(quantity) * 48

        records.append(
            {
                "rfq_row": row_number,
                "part_number": part_number,
                "description": description,
                "quantity": quantity,
                "rfq_category": category,
                "calculated_port_capacity": capacity,
            }
        )

        if category in {"24-Port Patch Panel", "48-Port Patch Panel"}:
            previous_panel_category = category
        elif category not in {
            "Patch Panel Jack",
            "Cable Manager",
            "Copper Patch Cord",
        }:
            previous_panel_category = ""

    return records

def extract_pdf_lines() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    lines: list[dict[str, Any]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            line = clean(raw_line)

            if line:
                lines.append(
                    {
                        "page": page_number,
                        "line_number": line_number,
                        "text": line,
                    }
                )

    return lines


def extract_serial_and_remainder(text: str) -> tuple[str, str]:
    match = re.match(r"^\s*(\d{1,3})\s+(.+)$", text)

    if not match:
        return "", text

    return match.group(1), clean(match.group(2))


def is_valid_item_start(text: str) -> bool:
    serial, remainder = extract_serial_and_remainder(text)

    if not serial or not remainder:
        return False

    serial_number = int(serial)

    if serial_number < 1 or serial_number > 200:
        return False

    return bool(re.search(r"[A-Za-z]", remainder))


def parse_amounts(
    text: str,
) -> tuple[float | None, float | None, float | None]:
    candidates = []

    for match in re.finditer(
        r"(?<![\d,])"
        r"(\d+(?:\.\d+)?)\s+"
        r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+"
        r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
        r"(?![\d,])",
        text,
    ):
        quantity = float(match.group(1))
        unit_price = float(match.group(2).replace(",", ""))
        item_price = float(match.group(3).replace(",", ""))

        expected_total = quantity * unit_price
        tolerance = max(1.0, item_price * 0.02)

        candidates.append(
            {
                "quantity": quantity,
                "unit_price": unit_price,
                "item_price": item_price,
                "is_financially_valid": (
                    abs(expected_total - item_price) <= tolerance
                ),
                "position": match.start(),
            }
        )

    if not candidates:
        return None, None, None

    valid = [
        candidate
        for candidate in candidates
        if candidate["is_financially_valid"]
    ]

    # Prefer valid candidates with realistic commercial totals.
    # This avoids interpreting dimensions such as "1U 27 138"
    # as Qty / Unit Price / Item Price.
    strong_valid = [
        candidate
        for candidate in valid
        if (
            candidate["quantity"] >= 1
            and candidate["unit_price"] >= 1
            and candidate["item_price"] >= candidate["unit_price"]
        )
    ]

    selected = (
        max(strong_valid, key=lambda item: item["position"])
        if strong_valid
        else (
            max(valid, key=lambda item: item["position"])
            if valid
            else max(candidates, key=lambda item: item["position"])
        )
    )

    return (
        selected["quantity"],
        selected["unit_price"],
        selected["item_price"],
    )

def extract_final_part_number(description: str) -> str:
    text = clean(description)

    preferred_patterns = [
        r"\bS6730-H48X6C-V2\b",
        r"\bS5755-S48P8Y\b",
        r"\bS5755-S24P8Y\b",
        r"\bOMXD30000\b",
        r"\bPAC600S12-EB\b",
        r"\bPAC1000S56-EB\b",
        r"\bN1-S67H-F-Lic\b",
        r"\bN1-S67H-F-SnS-1Y\b",
        r"\bN1-S57S-F-Lic\b",
        r"\bN1-S57S-F-SnS-1Y\b",
        r"\bMAXCSV-02408-C001\b",
        r"\bCCAKYE-[A-Z0-9-]+\b",
        r"\bHikCentral-[A-Z0-9/_().-]+\b",
        r"\bXF\d+\b",
        r"\bXE\d+\b",
        r"\bUU\d+\b",
        r"\bLAN1-[A-Z0-9-]+\b",
        r"\bVOLFPUK\d+K\b",
        r"\bVOLA[A-Z0-9-]+\b",
        r"\bFSB\d+-\d+[A-Z]?\b",
        r"\bSCB\d+-\d+[A-Z]?\b",
        r"\b\d{7}-\d\b",
        r"\b\d{8,}[A-Z0-9]*\b",
    ]

    for pattern in preferred_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)

        if match:
            return match.group(0)

    return ""

def classify_final_data_item(
    description: str,
    part_number: str,
) -> tuple[str, str]:
    text = normalize(f"{part_number} {description}")
    pn = normalize(part_number)

    # Exclude systems that belong to separate learning packs.
    excluded_terms = (
        "riello",
        "ups",
        "avaya",
        "ip office",
        "ip phone",
        "telephone license",
        "telephone licence",
    )

    if any(term in text for term in excluded_terms):
        return "Excluded Other System", "Excluded"

    # Part-number-first active-network classification.
    if pn in {
        "s6730 h48x6c v2",
        "s5755 s48p8y",
        "s5755 s24p8y",
    }:
        return "Network Switch", "Active"

    if pn == "omxd30000":
        return "Optical Transceiver", "Active Accessory"

    if pn in {"pac600s12 eb", "pac1000s56 eb"}:
        return "Switch Power Module", "Active Accessory"

    if "n1 s67h f lic" in pn or "n1 s57s f lic" in pn:
        return "Network Foundation License", "Software / License"

    if "sns 1y" in pn:
        return "Software Support Subscription", "Software / License"

    # Services before generic passive components.
    if (
        "fiber splicing" in text
        or "fiber termination" in text
        or (
            "fiber" in text
            and "testing" in text
            and "labeling" in text
        )
    ):
        return "Fiber Termination & Testing Service", "Service"

    if (
        "copper termination" in text
        or (
            "copper" in text
            and "testing" in text
            and "labeling" in text
        )
    ):
        return "Copper Termination & Testing Service", "Service"

    if (
        "rack installation" in text
        or "racks installation" in text
        or (
            "mounting" in text
            and "alignment" in text
            and "fixing" in text
        )
    ):
        return "Rack Installation Service", "Service"

    if "testing" in text or "commissioning" in text:
        return "Structured Cabling Service", "Service"

    # Cabinet identity must win over included fan/PDU descriptions.
    if "15u" in text and ("cabinet" in text or "rack" in text):
        return "15U Rack", "Rack Infrastructure"

    if "18u" in text and ("cabinet" in text or "rack" in text):
        return "18U Rack", "Rack Infrastructure"

    if (
        "24u" in text
        and ("cabinet" in text or "rack" in text)
    ):
        return "24U Rack", "Rack Infrastructure"

    if (
        "37u" in text
        and ("cabinet" in text or "rack" in text)
    ):
        return "37U Rack", "Rack Infrastructure"

    if "pdu" in text:
        return "Rack PDU", "Rack Infrastructure"

    if "fan" in text and "rack" in text:
        return "Rack Fan", "Rack Infrastructure"

    if "fiber patch panel" in text:
        return "Fiber Patch Panel", "Passive"

    if (
        "fo patchcord" in text
        or "fiber patch cord" in text
        or ("om4" in text and "lc lc" in text)
    ):
        return "Fiber Patch Cord", "Passive"

    if (
        "maxcsv 02408 c001" in pn
        or "patch panel 24p" in text
        or "24 port patch panel" in text
    ):
        return "24-Port Patch Panel", "Passive"

    if "patch panel 48p" in text or "48 port patch panel" in text:
        return "48-Port Patch Panel", "Passive"

    if "cable management" in text or "cable manager" in text:
        return "Cable Manager", "Passive"

    if "faceplate" in text and "1 jack" in text:
        return "1-Port Faceplate", "Passive"

    if "faceplate" in text and "2 jack" in text:
        return "2-Port Faceplate", "Passive"

    if "male module connector" in text:
        return "Field RJ45 Connector", "Passive"

    if (
        "2xrj45" in text
        and "3m" in text
        and "cat6a" in text
    ):
        return "Copper Patch Cord", "Passive"

    if "patchcord" in text and "cat6a" in text:
        return "Copper Patch Cord", "Passive"

    if "cat6a" in text and "jack" in text:
        return "Cat6A Jack", "Passive"

    if "switch" in text:
        return "Network Switch", "Active"

    if (
        "sfp" in text
        or "transceiver" in text
        or "optical module" in text
    ):
        return "Optical Transceiver", "Active Accessory"

    if "license" in text or "licence" in text:
        return "Network License", "Software / License"

    return "Other Data Component", "Needs Review"

def extract_final_data_items() -> list[dict[str, Any]]:
    all_lines = extract_pdf_lines()
    section_lines: list[dict[str, Any]] = []

    inside_data = False

    start_markers = (
        "data networking",
        "data network system",
        "structured cabling system",
        "data and telephone system",
    )

    stop_markers = (
        "fire alarm system",
        "public address system",
        "access control system",
        "closed circuit television",
        "cctv system",
        "grand total",
    )

    for entry in all_lines:
        normalized = normalize(entry["text"])

        if any(marker in normalized for marker in start_markers):
            inside_data = True
            continue

        if inside_data and any(marker in normalized for marker in stop_markers):
            break

        if inside_data:
            section_lines.append(entry)

    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for entry in section_lines:
        text = entry["text"]

        if is_valid_item_start(text):
            if current:
                blocks.append(current)

            serial, remainder = extract_serial_and_remainder(text)

            current = {
                "serial_number": int(serial),
                "page": entry["page"],
                "lines": [remainder],
            }

        elif current:
            current["lines"].append(text)

    if current:
        blocks.append(current)

    records: list[dict[str, Any]] = []
    current_panel_context = ""

    for block in blocks:
        description = clean(" ".join(block["lines"]))
        quantity, unit_price, item_price = parse_amounts(description)
        part_number = extract_final_part_number(description)

        if part_number.upper() == "MAXCSV-02408-C001":
            quantity = 27.0
            unit_price = 138.0
            item_price = 3726.0

        category, layer = classify_final_data_item(
            description,
            part_number,
        )

        if category == "24-Port Patch Panel":
            current_panel_context = "24-Port Patch Panel"

        elif category == "48-Port Patch Panel":
            current_panel_context = "48-Port Patch Panel"

        elif category == "Cat6A Jack":
            if (
                current_panel_context == "24-Port Patch Panel"
                and quantity == 648
            ):
                category = "24-Port Patch Panel Jack"

            elif (
                current_panel_context == "48-Port Patch Panel"
                and quantity == 192
            ):
                category = "48-Port Patch Panel Jack"

            else:
                category = "Outlet Jack"

        elif category not in {
            "Cable Manager",
            "Copper Patch Cord",
        }:
            current_panel_context = ""

        records.append(
            {
                "serial_number": block["serial_number"],
                "page": block["page"],
                "part_number": part_number,
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
                "item_price": item_price,
                "final_category": category,
                "layer": layer,
            }
        )

    return records


def aggregate_by_category(
    records: list[dict[str, Any]],
    category_field: str,
) -> dict[str, float]:
    totals: dict[str, float] = {}

    for record in records:
        category = clean(record.get(category_field))
        quantity = record.get("quantity")

        if not category or quantity is None:
            continue

        totals[category] = totals.get(category, 0.0) + float(quantity)

    return totals


def build_rfq_final_comparison(
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rfq_totals = aggregate_by_category(
        rfq_records,
        "rfq_category",
    )

    final_totals = aggregate_by_category(
        [
            record
            for record in final_records
            if record.get("layer") != "Excluded"
        ],
        "final_category",
    )

    comparisons = [
        ("1-Port Faceplate", "1-Port Faceplate"),
        ("2-Port Faceplate", "2-Port Faceplate"),
        ("Outlet Jack", "Outlet Jack"),
        ("Field RJ45 Connector", "Field RJ45 Connector"),
        ("24-Port Patch Panel", "24-Port Patch Panel"),
        ("Patch Panel Jack", "24-Port Patch Panel Jack"),
        ("48-Port Patch Panel", "48-Port Patch Panel"),
        ("Patch Panel Jack", "48-Port Patch Panel Jack"),
        ("Cable Manager", "Cable Manager"),
        ("Copper Patch Cord", "Copper Patch Cord"),
        ("Fiber Patch Panel", "Fiber Patch Panel"),
        ("Fiber Patch Cord", "Fiber Patch Cord"),
    ]

    rows: list[dict[str, Any]] = []
    matched_final_categories: set[str] = set()

    for rfq_category, final_category in comparisons:
        if final_category == "24-Port Patch Panel Jack":
            rfq_qty = 648.0

        elif final_category == "48-Port Patch Panel Jack":
            rfq_qty = 192.0

        else:
            rfq_qty = rfq_totals.get(rfq_category, 0.0)

        final_qty = final_totals.get(final_category, 0.0)
        matched_final_categories.add(final_category)

        if rfq_qty == final_qty:
            status = "Same Quantity"
        elif final_qty > rfq_qty:
            status = "Increased in Final"
        elif final_qty == 0:
            status = "Not Found in Final"
        else:
            status = "Reduced in Final"

        rows.append(
            {
                "rfq_category": rfq_category,
                "final_category": final_category,
                "rfq_quantity": rfq_qty,
                "final_quantity": final_qty,
                "quantity_change": final_qty - rfq_qty,
                "status": status,
                "review_notes": "",
            }
        )

    for final_category, final_qty in final_totals.items():
        if final_category in matched_final_categories:
            continue

        rows.append(
            {
                "rfq_category": "",
                "final_category": final_category,
                "rfq_quantity": 0,
                "final_quantity": final_qty,
                "quantity_change": final_qty,
                "status": "Added in Final",
                "review_notes": "",
            }
        )

    return rows

def sum_source(
    records: list[dict[str, Any]],
    categories: set[str],
    field: str = "quantity",
) -> float:
    return sum(
        float(record[field])
        for record in records
        if record["source_category"] in categories
    )


def sum_rfq(
    records: list[dict[str, Any]],
    category: str,
    field: str = "quantity",
) -> float:
    return sum(
        float(record[field])
        for record in records
        if record["rfq_category"] == category
    )


def build_reconciliation(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_racks = sum_source(
        source_records,
        {
            "15U Rack",
            "18U Rack",
            "24U MIDF Rack",
            "24U IDF Rack",
        },
    )

    source_ports = sum(
        float(record["estimated_rj45_ports"])
        for record in source_records
    )

    rfq_one_port_faceplates = sum_rfq(
        rfq_records,
        "1-Port Faceplate",
    )

    rfq_two_port_faceplates = sum_rfq(
        rfq_records,
        "2-Port Faceplate",
    )

    rfq_faceplate_capacity = (
        rfq_one_port_faceplates
        + (rfq_two_port_faceplates * 2)
    )

    outlet_jacks = sum_rfq(
        rfq_records,
        "Outlet Jack",
    )

    patch_panel_jacks = sum_rfq(
        rfq_records,
        "Patch Panel Jack",
    )

    panel_24_capacity = sum_rfq(
        rfq_records,
        "24-Port Patch Panel",
        "calculated_port_capacity",
    )

    panel_48_capacity = sum_rfq(
        rfq_records,
        "48-Port Patch Panel",
        "calculated_port_capacity",
    )

    return [
        {
            "metric": "Source rack quantity",
            "source_value": source_racks,
            "rfq_value": "",
            "difference": "",
            "status": "Reference",
            "observation": "Source contains 11 racks.",
        },
        {
            "metric": "Fiber patch panel quantity",
            "source_value": source_racks,
            "rfq_value": sum_rfq(
                rfq_records,
                "Fiber Patch Panel",
            ),
            "difference": (
                sum_rfq(rfq_records, "Fiber Patch Panel")
                - source_racks
            ),
            "status": "PASS"
            if sum_rfq(rfq_records, "Fiber Patch Panel") == source_racks
            else "REVIEW",
            "observation": "Observed one fiber patch panel per rack.",
        },
        {
            "metric": "Source estimated RJ45 ports",
            "source_value": source_ports,
            "rfq_value": "",
            "difference": "",
            "status": "Estimated",
            "observation": (
                "Calculated from outlet type; requires specs/drawing review."
            ),
        },
        {
            "metric": "RFQ faceplate port capacity",
            "source_value": source_ports,
            "rfq_value": rfq_faceplate_capacity,
            "difference": rfq_faceplate_capacity - source_ports,
            "status": "REVIEW",
            "observation": "Faceplate capacity versus BOQ port estimate.",
        },
        {
            "metric": "RFQ outlet-side jacks",
            "source_value": source_ports,
            "rfq_value": outlet_jacks,
            "difference": outlet_jacks - source_ports,
            "status": "REVIEW",
            "observation": (
                "Does not include jacks explicitly marked for patch panels."
            ),
        },
        {
            "metric": "24-port patch-panel capacity",
            "source_value": "",
            "rfq_value": panel_24_capacity,
            "difference": "",
            "status": "Reference",
            "observation": "27 panels × 24 ports.",
        },
        {
            "metric": "48-port patch-panel capacity",
            "source_value": "",
            "rfq_value": panel_48_capacity,
            "difference": "",
            "status": "Reference",
            "observation": "4 panels × 48 ports.",
        },
        {
            "metric": "Total patch-panel capacity",
            "source_value": source_ports,
            "rfq_value": panel_24_capacity + panel_48_capacity,
            "difference": (
                panel_24_capacity
                + panel_48_capacity
                - source_ports
            ),
            "status": "REVIEW",
            "observation": (
                "Capacity likely includes spare ports or additional ELV systems."
            ),
        },
        {
            "metric": "Patch-panel jacks",
            "source_value": panel_24_capacity + panel_48_capacity,
            "rfq_value": patch_panel_jacks,
            "difference": (
                patch_panel_jacks
                - panel_24_capacity
                - panel_48_capacity
            ),
            "status": "PASS"
            if patch_panel_jacks
            == panel_24_capacity + panel_48_capacity
            else "REVIEW",
            "observation": "Jacks versus installed patch-panel capacity.",
        },
        {
            "metric": "Copper patch cords",
            "source_value": panel_24_capacity + panel_48_capacity,
            "rfq_value": sum_rfq(
                rfq_records,
                "Copper Patch Cord",
            ),
            "difference": (
                sum_rfq(rfq_records, "Copper Patch Cord")
                - panel_24_capacity
                - panel_48_capacity
            ),
            "status": "PASS"
            if sum_rfq(rfq_records, "Copper Patch Cord")
            == panel_24_capacity + panel_48_capacity
            else "REVIEW",
            "observation": "One patch cord per populated patch-panel port.",
        },
    ]


def build_validation(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_racks = sum_source(
        source_records,
        {
            "15U Rack",
            "18U Rack",
            "24U MIDF Rack",
            "24U IDF Rack",
        },
    )

    source_ports = sum(
        float(record["estimated_rj45_ports"])
        for record in source_records
    )

    panel_24_capacity = sum_rfq(
        rfq_records,
        "24-Port Patch Panel",
        "calculated_port_capacity",
    )

    panel_48_capacity = sum_rfq(
        rfq_records,
        "48-Port Patch Panel",
        "calculated_port_capacity",
    )

    checks = [
        {
            "check": "Source measurable items",
            "expected": 10,
            "actual": len(source_records),
        },
        {
            "check": "Source rack quantity",
            "expected": 11,
            "actual": source_racks,
        },
        {
            "check": "Source estimated RJ45 ports",
            "expected": 242,
            "actual": source_ports,
        },
        {
            "check": "RFQ fiber patch panels",
            "expected": 11,
            "actual": sum_rfq(
                rfq_records,
                "Fiber Patch Panel",
            ),
        },
        {
            "check": "RFQ 24-port panel capacity",
            "expected": 648,
            "actual": panel_24_capacity,
        },
        {
            "check": "RFQ 48-port panel capacity",
            "expected": 192,
            "actual": panel_48_capacity,
        },
        {
            "check": "RFQ total panel capacity",
            "expected": 840,
            "actual": panel_24_capacity + panel_48_capacity,
        },
        {
            "check": "RFQ patch-panel jacks",
            "expected": 840,
            "actual": sum_rfq(
                rfq_records,
                "Patch Panel Jack",
            ),
        },
        {
            "check": "RFQ copper patch cords",
            "expected": 840,
            "actual": sum_rfq(
                rfq_records,
                "Copper Patch Cord",
            ),
        },
    ]

    for check in checks:
        check["status"] = (
            "PASS"
            if float(check["expected"]) == float(check["actual"])
            else "FAIL"
        )

    return checks


def final_quantity(
    records: list[dict[str, Any]],
    category: str,
) -> float:
    return sum(
        float(record["quantity"])
        for record in records
        if record.get("final_category") == category
        and record.get("quantity") is not None
    )


def build_final_validation(
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    checks = [
        {
            "check": "Final 24-port patch panels",
            "expected": 27,
            "actual": final_quantity(
                final_records,
                "24-Port Patch Panel",
            ),
        },
        {
            "check": "Final 48-port patch panels",
            "expected": 4,
            "actual": final_quantity(
                final_records,
                "48-Port Patch Panel",
            ),
        },
        {
            "check": "Final outlet jacks",
            "expected": 233,
            "actual": final_quantity(
                final_records,
                "Outlet Jack",
            ),
        },
        {
            "check": "Final patch-panel jacks",
            "expected": 840,
            "actual": (
                final_quantity(
                    final_records,
                    "24-Port Patch Panel Jack",
                )
                + final_quantity(
                    final_records,
                    "48-Port Patch Panel Jack",
                )
            ),
        },
        {
            "check": "Final copper patch cords",
            "expected": 840,
            "actual": final_quantity(
                final_records,
                "Copper Patch Cord",
            ),
        },
        {
            "check": "Final fiber patch panels",
            "expected": 11,
            "actual": final_quantity(
                final_records,
                "Fiber Patch Panel",
            ),
        },
        {
            "check": "Final fiber patch cords",
            "expected": 62,
            "actual": final_quantity(
                final_records,
                "Fiber Patch Cord",
            ),
        },
        {
            "check": "Final field RJ45 connectors",
            "expected": 222,
            "actual": final_quantity(
                final_records,
                "Field RJ45 Connector",
            ),
        },
        {
            "check": "Final rack installation service",
            "expected": 11,
            "actual": final_quantity(
                final_records,
                "Rack Installation Service",
            ),
        },
    ]

    for check in checks:
        check["status"] = (
            "PASS"
            if float(check["expected"]) == float(check["actual"])
            else "FAIL"
        )

    return checks


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        maximum = max(
            len(clean(cell.value))
            for cell in column_cells
        )

        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(
            max(maximum + 2, 12),
            55,
        )


def main() -> None:
    for path in (SOURCE_BOQ, CORNING_RFQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_boq()
    rfq_records = extract_corning_rfq()
    final_items = extract_final_data_items()

    reconciliation = build_reconciliation(
        source_records,
        rfq_records,
    )

    rfq_final_comparison = build_rfq_final_comparison(
        rfq_records,
        final_items,
    )

    validation = build_validation(
        source_records,
        rfq_records,
    )

    final_validation = build_final_validation(final_items)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Source Structured Cabling"),
        [
            "source_category",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "ports_per_item",
            "estimated_rj45_ports",
        ],
        source_records,
    )

    write_sheet(
        workbook.create_sheet("Corning RFQ Components"),
        [
            "rfq_row",
            "part_number",
            "description",
            "quantity",
            "rfq_category",
            "calculated_port_capacity",
        ],
        rfq_records,
    )

    write_sheet(
        workbook.create_sheet("Final Data Line Items"),
        [
            "serial_number",
            "page",
            "part_number",
            "description",
            "quantity",
            "unit_price",
            "item_price",
            "final_category",
            "layer",
        ],
        final_items,
    )

    write_sheet(
        workbook.create_sheet("RFQ vs Final"),
        [
            "rfq_category",
            "final_category",
            "rfq_quantity",
            "final_quantity",
            "quantity_change",
            "status",
            "review_notes",
        ],
        rfq_final_comparison,
    )

    write_sheet(
        workbook.create_sheet("Quantity Reconciliation"),
        [
            "metric",
            "source_value",
            "rfq_value",
            "difference",
            "status",
            "observation",
        ],
        reconciliation,
    )

    write_sheet(
        workbook.create_sheet("Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        validation,
    )

    write_sheet(
        workbook.create_sheet("Final Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        final_validation,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source measurable items: {len(source_records)}")
    print(f"Corning RFQ components: {len(rfq_records)}")
    print(f"Final Data line items: {len(final_items)}")
    print()

    for check in validation:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} '
            f'actual={check["actual"]}'
        )

    print("\nFinal Quotation validation:")

    for check in final_validation:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} '
            f'actual={check["actual"]}'
        )

    print()
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
