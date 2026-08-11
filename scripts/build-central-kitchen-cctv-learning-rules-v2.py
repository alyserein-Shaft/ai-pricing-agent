from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    "outputs/central-kitchen-cctv-learning-rules/"
    "Central_Kitchen_CCTV_Learning_Rules_v1.xlsx"
)

OUTPUT_FILE = Path(
    "outputs/central-kitchen-cctv-learning-rules/"
    "Central_Kitchen_CCTV_Learning_Rules_v2.xlsx"
)


RULE_GOVERNANCE = {
    "CCTV-CK-001": {
        "evidence_level": "Observed Once",
        "confidence_scope": "Single Historical Project",
        "generalization_status": "Project-Derived Candidate Rule",
        "vendor_dependency": "Product and environment dependent",
        "automation_level": "Discovery recommendation only",
        "approval_required": "Technical Engineer",
        "status": "Validated on 1 Historical Project",
        "quantity_relationship": (
            "Indoor Dome Qty + Outdoor Dome Qty "
            "= Consolidated Dome Product Qty"
        ),
        "consolidation_group_id": "CK-CCTV-DOME-001",
    },
    "CCTV-CK-002": {
        "evidence_level": "Observed Once",
        "confidence_scope": "Single Historical Project",
        "generalization_status": "Candidate Accessory Pattern",
        "vendor_dependency": "Accessory compatibility dependent",
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "status": "Validated on 1 Historical Project",
        "quantity_relationship": (
            "Camera Qty = BOQ Qty; Junction Box Qty = Camera Qty"
        ),
        "consolidation_group_id": "CK-CCTV-BULLET-001",
    },
    "CCTV-CK-003": {
        "evidence_level": "Observed Once",
        "confidence_scope": "Single Historical Project",
        "generalization_status": "Project-Derived Candidate Rule",
        "vendor_dependency": "Camera and mounting accessory dependent",
        "automation_level": "Discovery recommendation only",
        "approval_required": "Technical Engineer",
        "status": "Validated on 1 Historical Project",
        "quantity_relationship": (
            "15 Standard Anti-fog Bullet + 5 Pole-mounted "
            "= 20 Camera Products; Pole Mount Qty = 5 only"
        ),
        "consolidation_group_id": "CK-CCTV-AF-BULLET-001",
    },
    "CCTV-CK-004": {
        "evidence_level": "Observed Once",
        "confidence_scope": "Single Historical Project",
        "generalization_status": "Candidate Accessory Pattern",
        "vendor_dependency": "Accessory compatibility dependent",
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "status": "Validated on 1 Historical Project",
        "quantity_relationship": (
            "Camera Qty = BOQ Qty; Junction Box Qty = Camera Qty"
        ),
        "consolidation_group_id": "CK-CCTV-AF-DOME-001",
    },
    "CCTV-CK-005": {
        "evidence_level": "Mixed: Formula + Observed Once",
        "confidence_scope": (
            "Storage formula is reusable; selected architecture is project-specific"
        ),
        "generalization_status": "Structural Engineering Rule",
        "vendor_dependency": "Architecture and vendor dependent",
        "automation_level": "Calculation with mandatory engineer approval",
        "approval_required": "Technical Engineer",
        "status": "Approved as Structural Rule Only",
        "quantity_relationship": (
            "HDD Qty must be recalculated from camera count, bitrate, "
            "retention, FPS, RAID and usable capacity"
        ),
        "consolidation_group_id": "CK-CCTV-RECORDING-001",
    },
    "CCTV-CK-006": {
        "evidence_level": "Observed Once",
        "confidence_scope": "Single Historical Project",
        "generalization_status": "Project-Specific Component Pattern",
        "vendor_dependency": "Operational architecture dependent",
        "automation_level": "Discovery recommendation only",
        "approval_required": "Technical Engineer",
        "status": "Validated on 1 Historical Project",
        "quantity_relationship": (
            "Workstation and monitor quantities require operator-room inputs"
        ),
        "consolidation_group_id": "CK-CCTV-OPERATOR-001",
    },
    "CCTV-CK-007": {
        "evidence_level": "Mixed: Formula + Vendor-Dependent",
        "confidence_scope": (
            "Channel count relationship is reusable; license structure is vendor-specific"
        ),
        "generalization_status": "Vendor-Dependent Structural Rule",
        "vendor_dependency": "VMS licensing model dependent",
        "automation_level": "Recommendation with license-model verification",
        "approval_required": "Technical Engineer",
        "status": "Approved for Recommendation Only",
        "quantity_relationship": (
            "Channel License Qty = Licensed Camera Channels; "
            "Base License Qty depends on vendor architecture"
        ),
        "consolidation_group_id": "CK-CCTV-VMS-001",
    },
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_sheet(workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    headers = [clean(cell.value) for cell in ws[1]]

    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {
            headers[index]: row[index]
            for index in range(len(headers))
        }

        if any(clean(value) for value in row):
            records.append(record)

    return records


def normalize_rules(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    normalized = []

    for record in records:
        rule_id = clean(record.get("rule_id"))
        governance = RULE_GOVERNANCE.get(rule_id, {})

        result = dict(record)
        result.update(governance)

        normalized.append(result)

    return normalized


def parse_part_number_status(part_numbers: Any) -> str:
    text = clean(part_numbers)

    if not text:
        return "No Part Number"

    suspicious_fragments = (
        "10.4GT/s",
        "(STD",
        "(O-STD",
    )

    if any(fragment.lower() in text.lower() for fragment in suspicious_fragments):
        return "Needs Normalization"

    return "Extracted — Pending Product Library Verification"


def evidence_allocation(
    source_item: str,
) -> tuple[str, str, str]:
    if source_item in {
        "Indoor Dome Camera",
        "Outdoor Dome Camera",
    }:
        return (
            "CK-CCTV-DOME-001",
            "Source quantities consolidated into one final product line",
            "Shared final product quantity: 132",
        )

    if source_item == "Outdoor Bullet Camera":
        return (
            "CK-CCTV-BULLET-001",
            "Direct product mapping plus dedicated accessory",
            "47 cameras; 47 junction boxes",
        )

    if source_item == "Anti-fog Bullet Camera":
        return (
            "CK-CCTV-AF-BULLET-001",
            "Combined with pole-mounted variant",
            "15 of the combined 20 camera products",
        )

    if source_item == "Anti-fog Bullet Camera with Pole":
        return (
            "CK-CCTV-AF-BULLET-001",
            "Combined camera product; pole accessory allocated separately",
            "5 of the combined 20 cameras; 5 pole mounts",
        )

    if source_item == "Anti-fog Dome Camera":
        return (
            "CK-CCTV-AF-DOME-001",
            "Direct product mapping plus dedicated accessory",
            "14 cameras; 14 junction boxes",
        )

    if source_item == "NVR":
        return (
            "CK-CCTV-RECORDING-001",
            "System-level one-to-many breakdown",
            (
                "1 NVR; 15 HDD; 1 server; 1 workstation; "
                "2 monitors; 213 channel licenses"
            ),
        )

    return ("", "Needs Engineer Review", "")


def normalize_evidence(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    normalized = []

    for record in records:
        source_item = clean(record.get("source_item"))

        (
            consolidation_group_id,
            quantity_relationship,
            allocated_final_quantity,
        ) = evidence_allocation(source_item)

        result = dict(record)
        result.update(
            {
                "evidence_level": "Observed Once",
                "confidence_scope": "Central Kitchen - Makkah only",
                "generalization_status": (
                    "Historical Evidence — Not Global Rule"
                ),
                "consolidation_group_id": consolidation_group_id,
                "quantity_relationship": quantity_relationship,
                "allocated_final_quantity": allocated_final_quantity,
                "part_number_parse_status": parse_part_number_status(
                    record.get("final_part_numbers")
                ),
                "engineer_review_required": "Yes",
            }
        )

        normalized.append(result)

    return normalized


def build_validation(
    rules: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rule_ids = {clean(record.get("rule_id")) for record in rules}

    checks = [
        {
            "check": "Learning rules",
            "expected": 7,
            "actual": len(rules),
        },
        {
            "check": "Evidence mapping rows",
            "expected": 7,
            "actual": len(evidence),
        },
        {
            "check": "Rules marked with evidence level",
            "expected": 7,
            "actual": sum(
                bool(clean(record.get("evidence_level")))
                for record in rules
            ),
        },
        {
            "check": "Rules marked with confidence scope",
            "expected": 7,
            "actual": sum(
                bool(clean(record.get("confidence_scope")))
                for record in rules
            ),
        },
        {
            "check": "Evidence rows requiring engineer review",
            "expected": 7,
            "actual": sum(
                record.get("engineer_review_required") == "Yes"
                for record in evidence
            ),
        },
        {
            "check": "Unique rule IDs",
            "expected": 7,
            "actual": len(rule_ids),
        },
        {
            "check": "Global auto-approved rules",
            "expected": 0,
            "actual": sum(
                "Approved for Automation"
                in clean(record.get("status"))
                for record in rules
            ),
        },
    ]

    for check in checks:
        check["status"] = (
            "PASS"
            if check["expected"] == check["actual"]
            else "FAIL"
        )

    return checks


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        maximum = max(
            len(clean(cell.value))
            for cell in column_cells
        )

        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(
            max(maximum + 2, 14),
            55,
        )


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Missing input file: {INPUT_FILE}")

    source_workbook = load_workbook(INPUT_FILE, data_only=True)

    rules = normalize_rules(
        read_sheet(source_workbook, "Approved Learning Rules")
    )

    evidence = normalize_evidence(
        read_sheet(source_workbook, "Evidence Mapping")
    )

    validation = build_validation(rules, evidence)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Candidate Learning Rules"),
        [
            "rule_id",
            "rule_name",
            "source_items",
            "trigger_conditions",
            "transformation_type",
            "output_components",
            "quantity_logic",
            "project_evidence",
            "evidence_level",
            "confidence_scope",
            "generalization_status",
            "vendor_dependency",
            "automation_level",
            "approval_required",
            "quantity_relationship",
            "consolidation_group_id",
            "safety_constraints",
            "status",
        ],
        rules,
    )

    write_sheet(
        workbook.create_sheet("Evidence Mapping"),
        [
            "source_item",
            "source_description",
            "source_quantity",
            "rfq_descriptions",
            "final_categories",
            "final_part_numbers",
            "final_quantities",
            "relationship_type",
            "linked_rule_ids",
            "evidence_source",
            "evidence_level",
            "confidence_scope",
            "generalization_status",
            "consolidation_group_id",
            "quantity_relationship",
            "allocated_final_quantity",
            "part_number_parse_status",
            "engineer_review_required",
        ],
        evidence,
    )

    write_sheet(
        workbook.create_sheet("Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        validation,
    )

    workbook.save(OUTPUT_FILE)

    for check in validation:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} '
            f'actual={check["actual"]}'
        )

    print()
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
