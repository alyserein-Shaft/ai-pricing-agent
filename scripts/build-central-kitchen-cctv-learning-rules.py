from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    "outputs/central-kitchen-cctv-comparison/"
    "Central_Kitchen_CCTV_Comparison_v8.xlsx"
)

OUTPUT_DIR = Path("outputs/central-kitchen-cctv-learning-rules")

OUTPUT_FILE = (
    OUTPUT_DIR
    / "Central_Kitchen_CCTV_Learning_Rules_v1.xlsx"
)


RULES = [
    {
        "rule_id": "CCTV-CK-001",
        "rule_name": "Consolidate indoor and outdoor dome cameras",
        "source_items": "Indoor Dome Camera\nOutdoor Dome Camera",
        "trigger_conditions": (
            "Both BOQ items require compatible IP dome camera characteristics "
            "and the selected product is technically suitable for both locations."
        ),
        "transformation_type": "Technical Consolidation",
        "output_components": (
            "One consolidated dome camera model\n"
            "One junction box per installed camera"
        ),
        "quantity_logic": (
            "Consolidated camera quantity = indoor dome quantity "
            "+ outdoor dome quantity\n"
            "Junction box quantity = consolidated camera quantity"
        ),
        "project_evidence": "32 + 100 = 132 dome cameras; 132 junction boxes",
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Do not consolidate unless environmental rating, mounting method, "
            "lens, resolution, IR range and ingress protection are compatible."
        ),
        "status": "Approved from project evidence",
    },
    {
        "rule_id": "CCTV-CK-002",
        "rule_name": "Map outdoor bullet cameras with dedicated junction boxes",
        "source_items": "Outdoor Bullet Camera",
        "trigger_conditions": (
            "BOQ item is an outdoor wall-mounted IP bullet camera."
        ),
        "transformation_type": "Product + Accessory",
        "output_components": (
            "Outdoor bullet camera\n"
            "Compatible junction box"
        ),
        "quantity_logic": (
            "Camera quantity = BOQ quantity\n"
            "Junction box quantity = camera quantity"
        ),
        "project_evidence": "47 cameras; 47 junction boxes",
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Accessory must be explicitly compatible with the selected "
            "camera model and mounting environment."
        ),
        "status": "Approved from project evidence",
    },
    {
        "rule_id": "CCTV-CK-003",
        "rule_name": "Consolidate anti-fog bullet variants",
        "source_items": (
            "Anti-fog Bullet Camera\n"
            "Anti-fog Bullet Camera with Pole"
        ),
        "trigger_conditions": (
            "Both BOQ items require the same anti-fog bullet camera "
            "performance, with pole mounting required for a subset."
        ),
        "transformation_type": "Technical Consolidation",
        "output_components": (
            "One anti-fog bullet camera model\n"
            "Pole mount accessories for pole-mounted subset"
        ),
        "quantity_logic": (
            "Anti-fog bullet camera quantity = standard anti-fog bullet "
            "quantity + pole-mounted anti-fog bullet quantity\n"
            "Pole mount quantity = pole-mounted BOQ quantity only"
        ),
        "project_evidence": (
            "15 + 5 = 20 anti-fog bullet cameras; 5 pole mounts"
        ),
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Do not merge if lens, anti-fog technology, enclosure, mounting "
            "or environmental requirements differ."
        ),
        "status": "Approved from project evidence",
    },
    {
        "rule_id": "CCTV-CK-004",
        "rule_name": "Map anti-fog dome camera with junction box",
        "source_items": "Anti-fog Dome Camera",
        "trigger_conditions": (
            "BOQ item is an anti-fog ceiling-mounted dome camera."
        ),
        "transformation_type": "Product + Accessory",
        "output_components": (
            "Anti-fog dome camera\n"
            "Compatible junction box"
        ),
        "quantity_logic": (
            "Camera quantity = BOQ quantity\n"
            "Junction box quantity = camera quantity"
        ),
        "project_evidence": "14 cameras; 14 junction boxes",
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Verify anti-fog capability, enclosure rating, mounting method "
            "and accessory compatibility."
        ),
        "status": "Approved from project evidence",
    },
    {
        "rule_id": "CCTV-CK-005",
        "rule_name": "Expand NVR BOQ item into recording and storage subsystem",
        "source_items": "NVR",
        "trigger_conditions": (
            "BOQ includes an NVR requirement with channel count, recording "
            "duration, frame rate or storage requirements."
        ),
        "transformation_type": "Technical Refinement + System Breakdown",
        "output_components": (
            "NVR\n"
            "Surveillance HDDs\n"
            "Server where required\n"
            "Operating system license"
        ),
        "quantity_logic": (
            "NVR quantity based on channel capacity and redundancy design\n"
            "HDD quantity calculated from camera count, bitrate, recording "
            "duration, frame rate, RAID and usable capacity\n"
            "Server and OS quantities based on architecture"
        ),
        "project_evidence": (
            "1 NVR; 15 × 10 TB HDD; 1 server; "
            "1 Windows Server license"
        ),
        "automation_level": "Calculation + recommendation",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Never reuse project-specific HDD quantity without recalculation. "
            "Storage must be calculated for the current project inputs."
        ),
        "status": "Approved as structural rule only",
    },
    {
        "rule_id": "CCTV-CK-006",
        "rule_name": "Add CCTV management workstation and monitors",
        "source_items": "NVR / CCTV System Scope",
        "trigger_conditions": (
            "Project requires operator monitoring, administration or "
            "centralized CCTV management."
        ),
        "transformation_type": "Engineer Added System Component",
        "output_components": (
            "Operator workstation\n"
            "Monitoring displays"
        ),
        "quantity_logic": (
            "Workstation and monitor quantities must follow the project "
            "operator-room and operational requirements."
        ),
        "project_evidence": "1 workstation; 2 monitors",
        "automation_level": "Discovery recommendation",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Do not add automatically when operator positions, control-room "
            "scope or display requirements are not confirmed."
        ),
        "status": "Approved from project evidence",
    },
    {
        "rule_id": "CCTV-CK-007",
        "rule_name": "Add VMS licenses and commissioning services",
        "source_items": "CCTV System Scope",
        "trigger_conditions": (
            "Selected CCTV architecture uses licensed VMS software and "
            "requires system configuration and commissioning."
        ),
        "transformation_type": "Software + Service Breakdown",
        "output_components": (
            "VMS base license\n"
            "Camera channel licenses\n"
            "Testing and commissioning"
        ),
        "quantity_logic": (
            "Channel license quantity = number of licensed camera channels\n"
            "Base license quantity follows selected VMS architecture\n"
            "Testing and commissioning treated as service scope"
        ),
        "project_evidence": (
            "1 VMS base license; 213 channel licenses; "
            "testing and commissioning"
        ),
        "automation_level": "Recommendation only",
        "approval_required": "Technical Engineer",
        "safety_constraints": (
            "Confirm vendor licensing model, included channels, server "
            "architecture and commercial license terms."
        ),
        "status": "Approved from project evidence",
    },
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_relationships() -> list[dict[str, Any]]:
    workbook = load_workbook(INPUT_FILE, data_only=True)

    if "Relationships" not in workbook.sheetnames:
        raise ValueError("Relationships sheet not found")

    ws = workbook["Relationships"]
    headers = [clean(cell.value) for cell in ws[1]]

    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {
            headers[index]: row[index]
            for index in range(len(headers))
        }

        if clean(record.get("source_item")):
            records.append(record)

    return records


def build_evidence_mapping(
    relationships: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    mapping = []

    for relationship in relationships:
        source_item = clean(relationship.get("source_item"))

        matching_rule_ids = []

        for rule in RULES:
            if source_item.lower() in rule["source_items"].lower():
                matching_rule_ids.append(rule["rule_id"])

        if source_item in {
            "Indoor Dome Camera",
            "Outdoor Dome Camera",
        }:
            matching_rule_ids = ["CCTV-CK-001"]

        elif source_item == "Outdoor Bullet Camera":
            matching_rule_ids = ["CCTV-CK-002"]

        elif source_item in {
            "Anti-fog Bullet Camera",
            "Anti-fog Bullet Camera with Pole",
        }:
            matching_rule_ids = ["CCTV-CK-003"]

        elif source_item == "Anti-fog Dome Camera":
            matching_rule_ids = ["CCTV-CK-004"]

        elif source_item == "NVR":
            matching_rule_ids = [
                "CCTV-CK-005",
                "CCTV-CK-006",
                "CCTV-CK-007",
            ]

        mapping.append(
            {
                "source_item": source_item,
                "source_description": relationship.get(
                    "source_description",
                    "",
                ),
                "source_quantity": relationship.get(
                    "source_quantity",
                    "",
                ),
                "rfq_descriptions": relationship.get(
                    "rfq_descriptions",
                    "",
                ),
                "final_categories": relationship.get(
                    "final_categories",
                    "",
                ),
                "final_part_numbers": relationship.get(
                    "final_part_numbers",
                    "",
                ),
                "final_quantities": relationship.get(
                    "final_quantities",
                    "",
                ),
                "relationship_type": relationship.get(
                    "relationship_type",
                    "",
                ),
                "linked_rule_ids": ", ".join(matching_rule_ids),
                "evidence_source": (
                    "Central Kitchen - Makkah / "
                    "Q1067-626-LCU"
                ),
            }
        )

    return mapping


def validate(
    relationships: list[dict[str, Any]],
    evidence_mapping: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    linked_rules = set()

    for record in evidence_mapping:
        for rule_id in clean(record["linked_rule_ids"]).split(","):
            rule_id = rule_id.strip()
            if rule_id:
                linked_rules.add(rule_id)

    expected_rule_ids = {rule["rule_id"] for rule in RULES}

    checks = [
        {
            "check": "Approved learning rules",
            "expected": 7,
            "actual": len(RULES),
        },
        {
            "check": "Source relationship rows",
            "expected": 7,
            "actual": len(relationships),
        },
        {
            "check": "Evidence mapping rows",
            "expected": 7,
            "actual": len(evidence_mapping),
        },
        {
            "check": "All rules linked to evidence",
            "expected": 7,
            "actual": len(linked_rules & expected_rule_ids),
        },
        {
            "check": "Unlinked approved rules",
            "expected": 0,
            "actual": len(expected_rule_ids - linked_rules),
        },
    ]

    for check in checks:
        check["status"] = (
            "PASS"
            if check["expected"] == check["actual"]
            else "FAIL"
        )

    return checks


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            max_length = max(
                max_length,
                len(clean(cell.value)),
            )

        letter = get_column_letter(column_cells[0].column)

        ws.column_dimensions[letter].width = min(
            max(max_length + 2, 14),
            60,
        )


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Missing comparison file: {INPUT_FILE}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    relationships = read_relationships()
    evidence_mapping = build_evidence_mapping(relationships)
    validation = validate(relationships, evidence_mapping)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Approved Learning Rules"),
        [
            "rule_id",
            "rule_name",
            "source_items",
            "trigger_conditions",
            "transformation_type",
            "output_components",
            "quantity_logic",
            "project_evidence",
            "automation_level",
            "approval_required",
            "safety_constraints",
            "status",
        ],
        RULES,
    )

    write_sheet(
        workbook.create_sheet("Evidence Mapping"),
        [
            "source_item",
            "source_description",
            "source_quantity",
            "rfq_descriptions",
            "final_categories",
            "final_part_numbers",
            "final_quantities",
            "relationship_type",
            "linked_rule_ids",
            "evidence_source",
        ],
        evidence_mapping,
    )

    write_sheet(
        workbook.create_sheet("Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        validation,
    )

    workbook.save(OUTPUT_FILE)

    for check in validation:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} '
            f'actual={check["actual"]}'
        )

    print()
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
