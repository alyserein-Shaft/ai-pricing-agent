from __future__ import annotations

import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
ENGINEER_RFQ = BASE / "cctv-engineer-rfq.xlsx"
FINAL_QUOTATION = (
    BASE
    / "final-quotation"
    / "Q1067-626-LCU- Central Kitchen - Makkah.pdf"
)

OUTPUT_DIR = Path("outputs/central-kitchen-cctv-comparison")
OUTPUT_FILE = OUTPUT_DIR / "Central_Kitchen_CCTV_Comparison.xlsx"


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()

    replacements = {
        "cctv": "closed circuit television",
        "camera": "camera",
        "cameras": "camera",
        "no.": "no",
        "nos.": "no",
        "qty": "quantity",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"[^a-z0-9.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def best_text_cell(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if isinstance(value, str) and len(clean(value)) >= 8
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    numbers = []

    for value in values:
        number = numeric(value)

        if number is not None:
            numbers.append(number)

    if not numbers:
        return None

    for number in reversed(numbers):
        if number >= 1:
            return number

    return numbers[-1]


def find_unit(values: list[Any]) -> str:
    known_units = {
        "no",
        "nos",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "m",
        "meter",
        "meters",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in known_units:
            return clean(value)

    return ""


def extract_source_cctv() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    records: list[dict[str, Any]] = []

    target_sheets = [
        sheet
        for sheet in workbook.sheetnames
        if "28" in sheet or "cctv" in sheet.lower()
    ]

    for sheet_name in target_sheets:
        ws = workbook[sheet_name]
        inside_cctv = False

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))
            joined_normalized = normalize(joined)

            if not joined:
                continue

            if any(
                phrase in joined_normalized
                for phrase in (
                    "closed circuit television",
                    "video surveillance",
                    "cctv system",
                    "surveillance system",
                )
            ):
                inside_cctv = True

            if inside_cctv and any(
                phrase in joined_normalized
                for phrase in (
                    "access control system",
                    "fire alarm system",
                    "public address system",
                    "data network",
                    "structured cabling",
                )
            ):
                if "closed circuit" not in joined_normalized:
                    inside_cctv = False

            if not inside_cctv:
                continue

            description = best_text_cell(values)
            quantity = find_quantity(values)
            unit = find_unit(values)

            item_number = ""
            for value in values[:5]:
                text = clean(value)

                if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
                    item_number = text
                    break

            if not description:
                continue

            records.append(
                {
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "item_number": item_number,
                    "description": description,
                    "quantity": quantity,
                    "unit": unit,
                    "raw_row": joined,
                }
            )

    return records


def extract_engineer_rfq() -> list[dict[str, Any]]:
    workbook = load_workbook(ENGINEER_RFQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))

            if not joined:
                continue

            description = best_text_cell(values)
            quantity = find_quantity(values)
            unit = find_unit(values)

            part_number = ""

            for value in values:
                text = clean(value)

                if (
                    len(text) >= 4
                    and re.search(r"[A-Za-z]", text)
                    and re.search(r"\d", text)
                    and " " not in text
                    and not text.startswith("=")
                ):
                    part_number = text
                    break

            if not description:
                continue

            records.append(
                {
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "part_number": part_number,
                    "description": description,
                    "quantity": quantity,
                    "unit": unit,
                    "raw_row": joined,
                }
            )

    return records


def extract_final_cctv() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    records: list[dict[str, Any]] = []
    inside_cctv = False

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for line_number, line in enumerate(text.splitlines(), start=1):
            line = clean(line)

            if not line:
                continue

            normalized = normalize(line)

            if any(
                phrase in normalized
                for phrase in (
                    "closed circuit television",
                    "cctv",
                    "video surveillance",
                )
            ):
                inside_cctv = True

            if inside_cctv and any(
                phrase in normalized
                for phrase in (
                    "access control",
                    "data networking",
                    "public address",
                    "fire alarm",
                )
            ):
                if not any(
                    phrase in normalized
                    for phrase in (
                        "closed circuit television",
                        "cctv",
                    )
                ):
                    inside_cctv = False

            if not inside_cctv:
                continue

            part_numbers = re.findall(
                r"\b[A-Z0-9]+(?:[-/][A-Z0-9]+){1,}\b",
                line,
                flags=re.IGNORECASE,
            )

            numbers = re.findall(r"\b\d+(?:,\d{3})*(?:\.\d+)?\b", line)

            records.append(
                {
                    "page": page_number,
                    "line_number": line_number,
                    "part_number": part_numbers[0] if part_numbers else "",
                    "description": line,
                    "numbers": ", ".join(numbers),
                }
            )

    return records


def similarity(left: str, right: str) -> float:
    left_normalized = normalize(left)
    right_normalized = normalize(right)

    if not left_normalized or not right_normalized:
        return 0.0

    left_tokens = set(left_normalized.split())
    right_tokens = set(right_normalized.split())

    token_score = len(left_tokens & right_tokens) / max(
        len(left_tokens | right_tokens),
        1,
    )

    sequence_score = SequenceMatcher(
        None,
        left_normalized,
        right_normalized,
    ).ratio()

    return round((token_score * 0.65) + (sequence_score * 0.35), 4)


def build_comparison(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    comparisons = []

    for source in source_records:
        best_rfq = None
        best_rfq_score = 0.0

        for candidate in rfq_records:
            score = similarity(
                source["description"],
                candidate["description"],
            )

            if score > best_rfq_score:
                best_rfq = candidate
                best_rfq_score = score

        best_final = None
        best_final_score = 0.0

        for candidate in final_records:
            score = similarity(
                source["description"],
                candidate["description"],
            )

            if score > best_final_score:
                best_final = candidate
                best_final_score = score

        source_qty = source.get("quantity")
        rfq_qty = best_rfq.get("quantity") if best_rfq else None

        if best_rfq_score < 0.18:
            relationship = "Not Found in Engineer RFQ"
        elif (
            source_qty is not None
            and rfq_qty is not None
            and source_qty != rfq_qty
        ):
            relationship = "Quantity Difference"
        elif best_rfq_score >= 0.70:
            relationship = "Same / Refined Item"
        elif best_rfq_score >= 0.35:
            relationship = "Possible Breakdown / Refined"
        else:
            relationship = "Needs Engineer Review"

        comparisons.append(
            {
                "source_item": source.get("item_number", ""),
                "source_description": source["description"],
                "source_quantity": source_qty,
                "source_unit": source.get("unit", ""),
                "rfq_part_number": (
                    best_rfq.get("part_number", "") if best_rfq else ""
                ),
                "rfq_description": (
                    best_rfq.get("description", "") if best_rfq else ""
                ),
                "rfq_quantity": rfq_qty,
                "rfq_unit": best_rfq.get("unit", "") if best_rfq else "",
                "rfq_similarity": best_rfq_score,
                "final_part_number": (
                    best_final.get("part_number", "") if best_final else ""
                ),
                "final_description": (
                    best_final.get("description", "") if best_final else ""
                ),
                "final_page": best_final.get("page", "") if best_final else "",
                "final_similarity": best_final_score,
                "relationship": relationship,
                "engineer_decision": "",
                "review_notes": "",
            }
        )

    return comparisons


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column

        for cell in column_cells:
            value = clean(cell.value)
            max_length = max(max_length, len(value))

            if cell.row > 1:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        width = min(max(max_length + 2, 12), 55)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def write_records(ws, headers: list[str], records: list[dict[str, Any]]) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    style_sheet(ws)


def main() -> None:
    for path in (SOURCE_BOQ, ENGINEER_RFQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input file: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_cctv()
    rfq_records = extract_engineer_rfq()
    final_records = extract_final_cctv()

    comparisons = build_comparison(
        source_records,
        rfq_records,
        final_records,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_records(
        workbook.create_sheet("Source CCTV BOQ"),
        [
            "sheet",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        source_records,
    )

    write_records(
        workbook.create_sheet("Engineer CCTV RFQ"),
        [
            "sheet",
            "source_row",
            "part_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        rfq_records,
    )

    write_records(
        workbook.create_sheet("Final Quotation CCTV"),
        [
            "page",
            "line_number",
            "part_number",
            "description",
            "numbers",
        ],
        final_records,
    )

    write_records(
        workbook.create_sheet("Comparison"),
        [
            "source_item",
            "source_description",
            "source_quantity",
            "source_unit",
            "rfq_part_number",
            "rfq_description",
            "rfq_quantity",
            "rfq_unit",
            "rfq_similarity",
            "final_part_number",
            "final_description",
            "final_page",
            "final_similarity",
            "relationship",
            "engineer_decision",
            "review_notes",
        ],
        comparisons,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source CCTV rows: {len(source_records)}")
    print(f"Engineer RFQ rows: {len(rfq_records)}")
    print(f"Final quotation CCTV lines: {len(final_records)}")
    print(f"Comparison rows: {len(comparisons)}")
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
