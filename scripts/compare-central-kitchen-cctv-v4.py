from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
ENGINEER_RFQ = BASE / "cctv-engineer-rfq.xlsx"
FINAL_QUOTATION = (
    BASE
    / "final-quotation"
    / "Q1067-626-LCU- Central Kitchen - Makkah.pdf"
)

OUTPUT_DIR = Path("outputs/central-kitchen-cctv-comparison")
OUTPUT_FILE = OUTPUT_DIR / "Central_Kitchen_CCTV_Comparison_v8.xlsx"


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()

    replacements = {
        "anti-fog": "anti fog",
        "antifog": "anti fog",
        "anti  fog": "anti fog",
        "defog": "anti fog",
        "dfog": "anti fog",
        "dom type": "dome type",
        "dom camera": "dome camera",
        "n.v.r": "nvr",
        "licence": "license",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)

    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def longest_text(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if isinstance(value, str) and len(clean(value)) >= 8
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    candidates: list[float] = []

    for value in values:
        number = numeric(value)

        if number is not None:
            candidates.append(number)

    if not candidates:
        return None

    for number in reversed(candidates):
        if number >= 1:
            return number

    return candidates[-1]


def find_unit(values: list[Any]) -> str:
    units = {
        "no",
        "nos",
        "no.",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in units:
            return clean(value)

    return ""


def find_item_number(values: list[Any]) -> str:
    for value in values[:6]:
        text = clean(value)

        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
            return text

    return ""


def classify_camera_description(description: str) -> str:
    text = normalize(description)

    has_indoor = "indoor" in text
    has_outdoor = "outdoor" in text
    has_dome = "dome" in text or re.search(r"\bdom\b", text) is not None
    has_bullet = "bullet" in text
    has_anti_fog = "anti fog" in text
    has_pole = "pole" in text
    has_nvr = (
        "network video recorder" in text
        or re.search(r"\bnvr\b", text) is not None
    )

    if has_anti_fog and has_bullet and has_pole:
        return "Anti-fog Bullet Camera with Pole"

    if has_anti_fog and "wall mounted" in text and has_pole:
        return "Anti-fog Bullet Camera with Pole"

    if has_anti_fog and has_bullet:
        return "Anti-fog Bullet Camera"

    if has_anti_fog and "wall mounted" in text:
        return "Anti-fog Bullet Camera"

    if has_anti_fog and has_dome:
        return "Anti-fog Dome Camera"

    if has_anti_fog and "ceiling mounted" in text:
        return "Anti-fog Dome Camera"

    if has_indoor and has_dome:
        return "Indoor Dome Camera"

    if has_outdoor and has_dome:
        return "Outdoor Dome Camera"

    if has_outdoor and has_bullet:
        return "Outdoor Bullet Camera"

    if has_nvr:
        return "NVR"

    return ""


def extract_source_cctv() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if "28" not in sheet_name and "cctv" not in sheet_name.lower():
            continue

        ws = workbook[sheet_name]
        inside_cctv = False

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))
            normalized = normalize(joined)

            if not joined:
                continue

            if any(
                phrase in normalized
                for phrase in (
                    "closed circuit television",
                    "cctv system",
                    "video surveillance",
                )
            ):
                inside_cctv = True

            if inside_cctv and any(
                phrase in normalized
                for phrase in (
                    "card access system",
                    "access control system",
                    "fire alarm system",
                    "public address system",
                    "data networking",
                    "structured cabling",
                )
            ):
                if "closed circuit television" not in normalized:
                    break

            if not inside_cctv:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)

            if not description or quantity is None:
                continue

            source_item = classify_camera_description(description)

            if not source_item:
                continue

            records.append(
                {
                    "source_item": source_item,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "item_number": find_item_number(values),
                    "description": description,
                    "quantity": quantity,
                    "unit": find_unit(values),
                    "raw_row": joined,
                }
            )

    return records


def extract_engineer_rfq() -> list[dict[str, Any]]:
    workbook = load_workbook(ENGINEER_RFQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))

            if not joined:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)

            if not description or quantity is None:
                continue

            mapped_source_item = classify_camera_description(description)
            normalized = normalize(description)

            if not mapped_source_item:
                if "workstation" in normalized:
                    mapped_source_item = "Workstation"
                elif "monitor" in normalized:
                    mapped_source_item = "Monitor"
                elif "server" in normalized:
                    mapped_source_item = "Server"
                else:
                    continue

            records.append(
                {
                    "mapped_source_item": mapped_source_item,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "part_number": "",
                    "description": description,
                    "quantity": quantity,
                    "unit": find_unit(values),
                    "raw_row": joined,
                }
            )

    return records


def extract_pdf_lines() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    lines: list[dict[str, Any]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            line = clean(raw_line)

            if line:
                lines.append(
                    {
                        "page": page_number,
                        "line_number": line_number,
                        "text": line,
                    }
                )

    return lines


def parse_amounts(text: str) -> tuple[float | None, float | None, float | None]:
    matches = list(
        re.finditer(
            r"(?<!\d)"
            r"(\d+(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
            r"(?:\s|$)",
            text,
        )
    )

    if not matches:
        return None, None, None

    match = matches[-1]

    return (
        float(match.group(1)),
        float(match.group(2).replace(",", "")),
        float(match.group(3).replace(",", "")),
    )


def extract_serial_and_remainder(text: str) -> tuple[str, str]:
    match = re.match(r"^\s*(\d{1,2})\s+(.+)$", text)

    if not match:
        return "", text

    serial = match.group(1)
    remainder = clean(match.group(2))

    return serial, remainder


def is_valid_item_start(text: str) -> bool:
    serial, remainder = extract_serial_and_remainder(text)

    if not serial or not remainder:
        return False

    serial_number = int(serial)

    if serial_number < 1 or serial_number > 50:
        return False

    normalized = normalize(remainder)

    invalid_measurement_starts = (
        "mp ",
        "tb ",
        "ch ",
        "channel ",
        "fps ",
        "inch ",
    )

    if any(normalized.startswith(marker) for marker in invalid_measurement_starts):
        return False

    return bool(
        re.search(r"[A-Za-z]", remainder)
        and (
            re.search(r"[A-Za-z].*\d|\d.*[A-Za-z]", remainder)
            or any(
                term in normalized
                for term in (
                    "testing",
                    "commissioning",
                    "configuration",
                    "license",
                    "server",
                    "workstation",
                    "monitor",
                )
            )
        )
    )


def split_joined_items(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for entry in lines:
        text = entry["text"]

        if is_valid_item_start(text):
            if current:
                items.append(current)

            serial, remainder = extract_serial_and_remainder(text)

            current = {
                "serial_number": int(serial),
                "page": entry["page"],
                "lines": [remainder],
            }

        elif current:
            current["lines"].append(text)

    if current:
        items.append(current)

    return items


def extract_part_number(description: str) -> str:
    text = clean(description)

    server_match = re.search(
        r"\bPER\d+[A-Z0-9]*\b",
        text,
        flags=re.IGNORECASE,
    )
    if server_match:
        return server_match.group(0)

    server_match = re.search(r"\\bPER\\d+[A-Z0-9]*\\b", text, re.IGNORECASE)
    if server_match:
        return server_match.group(0)

    first_segment = text[:180]

    candidates = re.findall(
        r"\b[A-Z0-9][A-Z0-9()_.]*"
        r"(?:[-/][A-Z0-9()_.]+)+\b",
        first_segment,
        flags=re.IGNORECASE,
    )

    rejected_patterns = (
        r"^\d+TB",
        r"^\d+MP",
        r"^\d+CH",
        r"^\d+C/\d+T",
        r"^\d+FPS",
        r"^\d+VDC",
        r"^Q1067",
        r"^\d+TB/\d+TB",
    )

    for candidate in candidates:
        if any(
            re.match(pattern, candidate, flags=re.IGNORECASE)
            for pattern in rejected_patterns
        ):
            continue

        if re.search(r"[A-Za-z]", candidate) and re.search(r"\d", candidate):
            return candidate

    compact_candidate = re.match(
        r"^([A-Za-z0-9][A-Za-z0-9()_.\-/]{3,})\s+",
        first_segment,
    )

    if compact_candidate:
        candidate = compact_candidate.group(1)

        if re.search(r"[A-Za-z]", candidate) and re.search(r"\d", candidate):
            return candidate

    return ""


def classify_final_item(description: str, part_number: str) -> str:
    text = normalize(f"{part_number} {description}")

    if (
        "ds100hkai" in text
        or "industry standard 3 5 inch" in text
        or "designed for 24x7 operation" in text
        or re.search(r"\\b\\d+\\s*tb\\b", text)
    ):
        return "Storage HDD"

    if "junction box" in text:
        return "Junction Box"

    if "pole mount" in text or "pole mounting" in text:
        return "Pole Mount"

    if (
        "hard disk" in text
        or re.search(r"\bhdd\b", text)
        or re.search(r"\b10tb\b", text)
    ):
        return "Storage HDD"

    if "anti fog" in text and "bullet" in text:
        return "Anti-fog Bullet Camera"

    if "anti fog" in text and "dome" in text:
        return "Anti-fog Dome Camera"

    if "bullet network camera" in text or "bullet camera" in text:
        return "Outdoor Bullet Camera"

    if "dome network camera" in text or "dome camera" in text:
        return "Dome Camera Consolidated"

    if "network video recorder" in text or re.search(r"\bnvr\b", text):
        return "NVR"

    if "workstation" in text:
        return "Workstation"

    if "monitor" in text:
        return "Monitor"

    if "windows server" in text and "license" in text:
        return "Windows Server License"

    if "server" in text and "license" not in text:
        return "Server"

    if (
        "1ch" in text
        or "1 channel" in text
        or "channel license" in text
        or "channel licence" in text
    ):
        return "VMS Channel License"

    if "license" in text:
        return "VMS Base License"

    if (
        "testing and commissioning" in text
        or "testing commissioning" in text
        or "configuration and commissioning" in text
    ):
        return "Testing & Commissioning"

    return "Other CCTV Component"


def extract_final_cctv() -> list[dict[str, Any]]:
    all_lines = extract_pdf_lines()

    section_lines: list[dict[str, Any]] = []
    inside_cctv = False

    start_markers = (
        "closed circuit television",
        "cctv system",
        "video surveillance",
    )

    stop_markers = (
        "access control system",
        "public address system",
        "fire alarm system",
        "data networking",
        "structured cabling",
    )

    for entry in all_lines:
        normalized = normalize(entry["text"])

        if any(marker in normalized for marker in start_markers):
            inside_cctv = True
            continue

        if inside_cctv and any(marker in normalized for marker in stop_markers):
            inside_cctv = False
            break

        if not inside_cctv:
            continue

        if any(
            marker in normalized
            for marker in (
                "payment condition",
                "payment terms",
                "delivery period",
                "warranty",
            )
        ):
            continue

        section_lines.append(entry)

    item_blocks = split_joined_items(section_lines)
    records: list[dict[str, Any]] = []

    for block in item_blocks:
        description = clean(" ".join(block["lines"]))

        quantity, unit_price, item_price = parse_amounts(description)
        part_number = extract_part_number(description)
        category = classify_final_item(description, part_number)

        records.append(
            {
                "serial_number": block["serial_number"],
                "page": block["page"],
                "mapped_category": category,
                "part_number": part_number,
                "description": description,
                "quantity": quantity,
                "unit_price": unit_price,
                "item_price": item_price,
            }
        )

    return records


def quantity_for_category(
    records: list[dict[str, Any]],
    category: str,
) -> float:
    return sum(
        float(record["quantity"])
        for record in records
        if record["mapped_category"] == category
        and record["quantity"] is not None
    )


def select_final_matches(
    source_name: str,
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_category: dict[str, list[dict[str, Any]]] = {}

    for record in final_records:
        by_category.setdefault(
            record["mapped_category"],
            [],
        ).append(record)

    selected: list[dict[str, Any]] = []

    if source_name in {"Indoor Dome Camera", "Outdoor Dome Camera"}:
        selected.extend(by_category.get("Dome Camera Consolidated", []))

        selected.extend(
            record
            for record in by_category.get("Junction Box", [])
            if record.get("quantity") == 132
        )

    elif source_name == "Outdoor Bullet Camera":
        selected.extend(by_category.get("Outdoor Bullet Camera", []))

        selected.extend(
            record
            for record in by_category.get("Junction Box", [])
            if record.get("quantity") == 47
        )

    elif source_name == "Anti-fog Bullet Camera":
        selected.extend(by_category.get("Anti-fog Bullet Camera", []))

    elif source_name == "Anti-fog Bullet Camera with Pole":
        selected.extend(by_category.get("Anti-fog Bullet Camera", []))

        selected.extend(
            record
            for record in by_category.get("Pole Mount", [])
            if record.get("quantity") == 5
        )

    elif source_name == "Anti-fog Dome Camera":
        selected.extend(by_category.get("Anti-fog Dome Camera", []))

        selected.extend(
            record
            for record in by_category.get("Junction Box", [])
            if record.get("quantity") == 14
        )

    elif source_name == "NVR":
        categories = {
            "NVR",
            "Storage HDD",
            "Server",
            "Windows Server License",
            "Workstation",
            "Monitor",
            "VMS Base License",
            "VMS Channel License",
            "Testing & Commissioning",
        }

        selected.extend(
            record
            for record in final_records
            if record["mapped_category"] in categories
        )

    return selected


def build_relationships(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rfq_by_name: dict[str, list[dict[str, Any]]] = {}

    for record in rfq_records:
        rfq_by_name.setdefault(
            record["mapped_source_item"],
            [],
        ).append(record)

    relationships: list[dict[str, Any]] = []

    for source in source_records:
        source_name = source["source_item"]
        rfq_matches = rfq_by_name.get(source_name, [])
        final_matches = select_final_matches(source_name, final_records)

        if source_name in {"Indoor Dome Camera", "Outdoor Dome Camera"}:
            relationship_type = "Technical Consolidation"

        elif source_name == "NVR":
            relationship_type = "Technical Refinement + System Breakdown"

        elif source_name == "Anti-fog Bullet Camera with Pole":
            relationship_type = "Consolidated Product + Pole Accessory"

        elif len(final_matches) > 1:
            relationship_type = "Product + Accessory"

        else:
            relationship_type = "Direct Product Mapping"

        relationships.append(
            {
                "source_item": source_name,
                "source_description": source["description"],
                "source_quantity": source["quantity"],
                "source_unit": source["unit"],
                "rfq_descriptions": "\n".join(
                    record["description"] for record in rfq_matches
                ),
                "rfq_quantities": ", ".join(
                    str(record["quantity"]) for record in rfq_matches
                ),
                "final_categories": "\n".join(
                    record["mapped_category"] for record in final_matches
                ),
                "final_part_numbers": "\n".join(
                    record["part_number"]
                    for record in final_matches
                    if record["part_number"]
                ),
                "final_quantities": ", ".join(
                    str(record["quantity"])
                    for record in final_matches
                    if record["quantity"] is not None
                ),
                "relationship_type": relationship_type,
                "engineering_observation": "",
                "approved_learning_rule": "",
            }
        )

    return relationships


def total_quantity(
    records: list[dict[str, Any]],
    names: set[str],
    field: str,
) -> float:
    return sum(
        float(record["quantity"])
        for record in records
        if record.get(field) in names
        and record.get("quantity") is not None
    )


def validate_results(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
    relationships: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_camera_names = {
        "Indoor Dome Camera",
        "Outdoor Dome Camera",
        "Outdoor Bullet Camera",
        "Anti-fog Bullet Camera",
        "Anti-fog Bullet Camera with Pole",
        "Anti-fog Dome Camera",
    }

    final_camera_categories = {
        "Dome Camera Consolidated",
        "Outdoor Bullet Camera",
        "Anti-fog Bullet Camera",
        "Anti-fog Dome Camera",
    }

    source_camera_qty = total_quantity(
        source_records,
        source_camera_names,
        "source_item",
    )

    rfq_camera_qty = total_quantity(
        rfq_records,
        source_camera_names,
        "mapped_source_item",
    )

    final_camera_qty = total_quantity(
        final_records,
        final_camera_categories,
        "mapped_category",
    )

    checks = [
        {
            "check": "Source measurable items",
            "expected": 7,
            "actual": len(source_records),
        },
        {
            "check": "Source camera quantity",
            "expected": 213,
            "actual": source_camera_qty,
        },
        {
            "check": "RFQ camera quantity",
            "expected": 213,
            "actual": rfq_camera_qty,
        },
        {
            "check": "Final camera quantity",
            "expected": 213,
            "actual": final_camera_qty,
        },
        {
            "check": "Relationship rows",
            "expected": 7,
            "actual": len(relationships),
        },
        {
            "check": "Final NVR quantity",
            "expected": 1,
            "actual": quantity_for_category(final_records, "NVR"),
        },
        {
            "check": "Final Workstation quantity",
            "expected": 1,
            "actual": quantity_for_category(final_records, "Workstation"),
        },
        {
            "check": "Final HDD quantity",
            "expected": 15,
            "actual": quantity_for_category(final_records, "Storage HDD"),
        },
        {
            "check": "Final Channel License quantity",
            "expected": 213,
            "actual": quantity_for_category(
                final_records,
                "VMS Channel License",
            ),
        },
    ]

    for check in checks:
        check["status"] = (
            "PASS"
            if float(check["actual"]) == float(check["expected"])
            else "FAIL"
        )

    return checks


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = max(
            len(clean(cell.value))
            for cell in column_cells
        )

        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(
            max(max_length + 2, 12),
            55,
        )


def main() -> None:
    for path in (SOURCE_BOQ, ENGINEER_RFQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_cctv()
    rfq_records = extract_engineer_rfq()
    final_records = extract_final_cctv()

    relationships = build_relationships(
        source_records,
        rfq_records,
        final_records,
    )

    checks = validate_results(
        source_records,
        rfq_records,
        final_records,
        relationships,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Source CCTV BOQ"),
        [
            "source_item",
            "sheet",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        source_records,
    )

    write_sheet(
        workbook.create_sheet("Engineer CCTV RFQ"),
        [
            "mapped_source_item",
            "sheet",
            "source_row",
            "part_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        rfq_records,
    )

    write_sheet(
        workbook.create_sheet("Final Quotation CCTV"),
        [
            "serial_number",
            "page",
            "mapped_category",
            "part_number",
            "description",
            "quantity",
            "unit_price",
            "item_price",
        ],
        final_records,
    )

    write_sheet(
        workbook.create_sheet("Relationships"),
        [
            "source_item",
            "source_description",
            "source_quantity",
            "source_unit",
            "rfq_descriptions",
            "rfq_quantities",
            "final_categories",
            "final_part_numbers",
            "final_quantities",
            "relationship_type",
            "engineering_observation",
            "approved_learning_rule",
        ],
        relationships,
    )

    write_sheet(
        workbook.create_sheet("Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        checks,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source CCTV measurable items: {len(source_records)}")
    print(f"Engineer RFQ measurable rows: {len(rfq_records)}")
    print(f"Final CCTV line items: {len(final_records)}")
    print(f"Relationship rows: {len(relationships)}")
    print()

    for check in checks:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} '
            f'actual={check["actual"]}'
        )

    print()
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
