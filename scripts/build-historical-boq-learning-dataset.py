from __future__ import annotations

import hashlib, json, re, sqlite3, sys
from collections import defaultdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path('/Users/serein-b/Downloads/projects')
DB = Path('.wrangler/state/v3/d1/miniflare-D1DatabaseObject/faaf2b0445ab934c3aac48ddf0cdfade8f9bac050be98993748742cdd2cb05fb.sqlite')
OUT = Path('outputs/historical-boq-learning')
ACTOR = 'local-development-user'
CHECKSUM_CACHE = Path('/tmp/five-project-files.sha256')
CHECKSUMS = {}
if CHECKSUM_CACHE.exists():
    for line in CHECKSUM_CACHE.read_text(errors='ignore').splitlines():
        if '  ' in line:
            checksum, file_path = line.split('  ', 1); CHECKSUMS[file_path] = checksum

PROJECTS = {
    'Bab Al khair - Makkah': ('Bab Al-Khair Hospital — Makkah', 'IHCC', ['Low Current', 'Fire Alarm', 'CCTV', 'Access Control', 'Data', 'UPS'], '2026-06-07', 'Partial Learning Pair'),
    'Dialysis Center Building -Makkah': ('Dialysis Center Building — Makkah', 'Sanabel Contracting Company', ['Low Current', 'Electrical'], '2026-04-12', 'Missing Source BOQ'),
    'Central Kitchen - Makkah': ('Central Kitchen — Makkah', 'Abdul Latif Jameel Group', ['Low Current', 'Fire Alarm', 'CCTV', 'Access Control', 'Data'], '2026-06-27', 'Partial Learning Pair'),
    'Opera Block Townhouses-Diriyah': ('Opera Block Townhouses — Diriyah', 'Becarabia Company', ['Fire Alarm', 'CCTV', 'Access Control', 'ICT', 'Audio Visual', 'Public Address', 'Intercom'], '2026-04-22', 'Partial Learning Pair'),
    'Construction of The BTS Multifamily Plots for MARAFY Commercial Core- ICT-Jeddah': ('BTS Multifamily Plots for MARAFY Commercial Core — ICT Jeddah', 'Mobco Group', ['Fire Alarm', 'CCTV', 'Access Control', 'Data', 'Disabled Toilet Alarm'], '2026-04-28', 'Complete Learning Pair'),
}

def stable(prefix: str, value: str) -> str:
    return f'{prefix}_{hashlib.sha256(value.encode()).hexdigest()[:24]}'

def sha(path: Path) -> str:
    if str(path) in CHECKSUMS: return CHECKSUMS[str(path)]
    h = hashlib.sha256()
    with path.open('rb') as f:
        while b := f.read(1024 * 1024): h.update(b)
    return h.hexdigest()

def norm(value) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(value or '').lower()).strip()

def j(value) -> str: return json.dumps(value, ensure_ascii=False, default=str)

def workbook_info(path: Path):
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        return [ws.title for ws in wb.worksheets], 'Readable'
    except Exception as e:
        return [], 'Encrypted / Needs Review' if 'zip' in str(e).lower() or 'password' in str(e).lower() else f'Unreadable: {type(e).__name__}'

def file_role(path: Path):
    s = str(path).lower(); n = path.name.lower(); ext = path.suffix.lower()
    if '/quotation/' in s and ext == '.pdf': return 'Final Quotation', 'Downstream Output', 95, 'Quotation-folder placement plus quotation reference and issued PDF; completion/finality still requires human confirmation.'
    if '/from supplier/' in s or 'supplier' in s: return 'Supplier Quotation', 'Supporting', 85, 'Supplier-folder placement; not treated as source BOQ.'
    if 'spec' in n or '/specs/' in s: return 'Specification', 'Supporting', 80, 'Specification naming/folder evidence.'
    if ext in ('.dwg', '.dwl', '.dwl2') or '/dwg/' in s or 'drawing' in s: return 'Drawing', 'Supporting', 90, 'CAD/drawing extension or drawing folder.'
    if 'vendor' in s or 'submittal' in s: return 'Technical Submittal', 'Supporting', 70, 'Vendor/submittal folder evidence; requires review.'
    if ext in ('.xlsx', '.xls'):
        if re.match(r'q\d+', n): return 'Final Cost Sheet', 'Downstream Output', 80, 'Quotation-number workbook; encrypted workbooks require human verification.'
        if 'npq' in n: return 'Supporting Document', 'Supporting', 80, 'New Project Qualification workbook.'
        if 'updated' in n or 'rev0' in s: return 'Revised BOQ', 'Source', 75, 'Revision/update evidence in path or filename.'
        if 'boq' in n or 'rfq' in n: return 'Original Client BOQ', 'Source', 70, 'BOQ/RFQ workbook naming; authority requires pair review.'
        return 'Unknown / Needs Review', 'Unknown', 30, 'Spreadsheet without reliable role evidence.'
    if ext == '.pdf': return 'Supporting Document', 'Supporting', 45, 'PDF without reliable role evidence.'
    return 'Unknown / Needs Review', 'Unknown', 20, 'No supported authoritative role evidence.'

def is_selected_source(project_folder: str, path: Path) -> bool:
    s = str(path); n = path.name.lower()
    if path.suffix.lower() not in ('.xlsx', '.xls') or re.match(r'q\d+', n) or 'npq' in n or '/From Supplier/' in s: return False
    if project_folder == 'Dialysis Center Building -Makkah': return False
    if project_folder == 'Central Kitchen - Makkah': return n == 'copy of electrical boq-for client use.xlsx'
    if project_folder == 'Construction of The BTS Multifamily Plots for MARAFY Commercial Core- ICT-Jeddah':
        return ('Rev02/Data From Client - Eslam' in s and n in ('2380 lc boq.xlsx','2384 lc boq.xlsx'))
    if project_folder == 'Opera Block Townhouses-Diriyah': return '0733-dgcl_opera block townhouses' in n
    if project_folder == 'Bab Al khair - Makkah': return path.parent.name == 'Data' and n == 'lc system boq. - bab al khair makkah project.xlsx'
    return False

def detect_columns(ws):
    aliases = {'item': ('item','s. no','s/n','item no'), 'description': ('description','descr iption','material'), 'quantity': ('quantity','qty','boq qty'), 'unit': ('unit','uom')}
    best = None
    for r in range(1, min(ws.max_row, 40) + 1):
        found = {}
        for c in range(1, min(ws.max_column, 30) + 1):
            v = norm(ws.cell(r,c).value)
            for key, names in aliases.items():
                if any(v == norm(x) or norm(x) in v for x in names): found.setdefault(key,c)
            # Prefer the workbook's explicit canonical headers over broader aliases
            # such as "Material", which may contain only a discipline/category.
            if v == 'description': found['description'] = c
            if v in ('quantity', 'qty', 'boq qty'): found['quantity'] = c
            if v in ('unit', 'uom'): found['unit'] = c
            if v in ('item', 'item no', 's no', 's n'): found['item'] = c
        score = len(found)
        if best is None or score > best[0]: best=(score,r,found)
    return best or (0,1,{})

def classify_row(values, desc, item, unit, qty):
    populated = [v for v in values if v not in (None,'')]
    text = norm(' '.join(str(v) for v in populated))
    if not populated: return 'Blank / Formatting', 'No populated cells.'
    if any(x in text for x in ('grand total','total amount')): return 'Total', 'Explicit total wording.'
    if 'subtotal' in text or 'carried to collection' in text: return 'Subtotal', 'Explicit subtotal/collection wording.'
    if any(x in text for x in ('notes','note ')) and qty in (None,''): return 'Note', 'Explicit note wording without measurable quantity.'
    if desc and qty not in (None,'') and unit not in (None,''): return 'BOQ Item', 'Description, unit and quantity are explicitly populated.'
    if desc and item and qty in (None,'') and len(str(desc)) < 140: return 'Section', 'Item/description hierarchy row without quantity.'
    if desc and qty in (None,'') and len(str(desc)) > 140: return 'Preamble', 'Long scope text without quantity.'
    return 'Needs Review', 'Populated row does not meet a deterministic row-type rule.'

def extract_source_rows(project_id, file_id, path: Path):
    rows=[]
    try: wb=load_workbook(path, read_only=True, data_only=False)
    except Exception: return rows
    for ws in wb.worksheets:
        _, header_row, cols = detect_columns(ws)
        merged = []
        for r, raw_row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column, 64), values_only=True), 1):
            # Some legacy workbooks carry formatting to Excel's last column. Bound physical extraction.
            values=list(raw_row)
            desc=values[cols.get('description',2)-1] if values else None
            item=values[cols.get('item',1)-1] if values else None
            unit=values[cols.get('unit',3)-1] if values else None
            qty=values[cols.get('quantity',4)-1] if values else None
            row_type, reason=classify_row(values,desc,item,unit,qty)
            formulas={str(c):v for c,v in enumerate(values,1) if isinstance(v,str) and v.startswith('=')}
            merge_context=[x for x in merged if re.search(rf'\b[A-Z]+{r}\b',x)]
            rid=stable('histRow',f'{file_id}|{ws.title}|{r}')
            rows.append({'id':rid,'project_id':project_id,'file_id':file_id,'sheet':ws.title,'row':r,'item':item,'section':None,'cells':values,'description':desc,'unit':unit,'quantity':qty,'formulae':formulas,'discipline':None,'system':None,'location':None,'formatting':{'headerRow':header_row,'streamingExtraction':True},'merged':merge_context,'hidden':False,'row_type':row_type,'reason':reason,'provenance':{'path':str(path),'sheet':ws.title,'row':r,'mergedAndHiddenState':'Not available in bounded streaming extraction; retained for human review.'}})
    return rows

def extract_final_rows(project_id, file_id, path: Path):
    out=[]
    try: reader=PdfReader(str(path))
    except Exception: return out
    for page_no,page in enumerate(reader.pages,1):
        text=page.extract_text() or ''
        lines=[re.sub(r'\s+',' ',x).strip() for x in text.splitlines() if x.strip()]
        for ix,line in enumerate(lines):
            # Preserve only lines that look like issued quotation detail, not narrative headings.
            m=re.match(r'^(?:Electrical\s+\d+\s+[^\d]+\s+)?([A-Z0-9.\-/]+)\s+(.{8,}?)\s+(No|m|Set|Lot|Item|LS)\s+([0-9,.]+)\s+([0-9,.]+|included|excluded)(?:\s+([0-9,.]+|included|excluded))?$',line,re.I)
            p=re.match(r'^(\d+)\s+([A-Z0-9][A-Z0-9.\-/]+)\s+(.{8,}?)\s+([0-9,.]+)\s+([0-9,.]+)\s+([0-9,.]+)$',line)
            if not (m or p): continue
            if m: item,desc,unit,qty=m.group(1),m.group(2),m.group(3),m.group(4); part=None
            else: item,part,desc,qty=p.group(1),p.group(2),p.group(3),p.group(4); unit='No'
            fid=stable('histFinalRow',f'{file_id}|{page_no}|{ix}|{line}')
            out.append({'id':fid,'project_id':project_id,'file_id':file_id,'page':page_no,'ref':item,'type':'BOQ Item','description':desc,'unit':unit,'quantity':qty,'discipline':None,'system':None,'category':None,'assembly':None,'split_merge':None,'exclusion':'Excluded' if 'excluded' in line.lower() else None,'manufacturer':None,'part_number':part,'accessories':None,'notes':None,'approval':'Issued Quotation / Needs Review','provenance':{'path':str(path),'page':page_no,'line':line}})
    return out

def align(project_name, sources, finals):
    alignments=[]; used=set()
    if 'MARAFY' not in project_name.upper(): return alignments
    for src in sources:
        if src['row_type']!='BOQ Item' or not src['description']: continue
        sd=norm(src['description']); candidates=[]
        for fin in finals:
            fd=norm(fin['description']); ratio=SequenceMatcher(None,sd,fd).ratio()
            if sd and (sd in fd or fd in sd): ratio=max(ratio,.96)
            if ratio>=.62: candidates.append((ratio,fin))
        if not candidates: continue
        ratio,fin=max(candidates,key=lambda x:x[0]); used.add(fin['id'])
        quantity_equal=norm(src['quantity'])==norm(fin['quantity']) and norm(src['quantity'])!=''
        unit_equal=norm(src['unit'])==norm(fin['unit']) and norm(src['unit'])!=''
        if ratio>=.96 and quantity_equal and unit_equal: outcome='Exact Alignment'; confidence=98; eligible=1
        elif ratio>=.88 and (quantity_equal or unit_equal): outcome='Strong Alignment'; confidence=90; eligible=1
        elif ratio>=.72: outcome='Possible Alignment'; confidence=70; eligible=0
        else: outcome='Unresolved'; confidence=55; eligible=0
        alignments.append({'id':stable('histAlignment',src['id']+'|'+fin['id']),'project_id':src['project_id'],'source_row_id':src['id'],'final_row_id':fin['id'],'outcome':outcome,'method':'Description + unit + quantity + sequence','confidence':confidence,'evidence':{'descriptionSimilarity':round(ratio,4),'quantityEqual':quantity_equal,'unitEqual':unit_equal},'eligible':eligible})
    return alignments

def main():
    OUT.mkdir(parents=True,exist_ok=True)
    conn=sqlite3.connect(DB); conn.execute('PRAGMA foreign_keys=ON')
    before={t:conn.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0] for t in ('projects','boq_items','product_match_runs','pricing_runs')}
    inventory=[]; source_rows=[]; final_rows=[]; alignments=[]; project_records=[]
    for folder,(name,client,disciplines,date,pair) in PROJECTS.items():
        pid=stable('historicalProject',folder); root=ROOT/folder
        files=[]
        for path in sorted(p for p in root.rglob('*') if p.is_file()):
            checksum=sha(path); role,side,confidence,evidence=file_role(path); sheets=[]; readability='Registered'
            if path.suffix.lower() in ('.xlsx','.xls'): readability='Registered / Sheet review pending'
            elif path.suffix.lower()=='.pdf': readability='Readable PDF'
            fid=stable('historicalFile',pid+'|'+str(path.relative_to(root)))
            row={'id':fid,'project_id':pid,'path':str(path),'name':path.name,'checksum':checksum,'size':path.stat().st_size,'extension':path.suffix.lower(),'sheets':sheets,'role':role,'side':side,'revision':next(iter(re.findall(r'(?i)(?:rev|r)[ _-]?(\d+)',path.name)),None),'evidence':evidence,'confidence':confidence,'human_review':1,'readability':readability}
            files.append(row); inventory.append(row)
        project_records.append({'id':pid,'name':name,'client':client,'disciplines':disciplines,'date':date,'status':'Completed Evidence Supplied / Needs Review','pair':pair,'root':str(root),'completion':{'issuedQuotation':any(f['role']=='Final Quotation' for f in files)}})
        seen_source_checksums=set()
        for f in files:
            path=Path(f['path'])
            if is_selected_source(folder,path) and f['checksum'] not in seen_source_checksums:
                seen_source_checksums.add(f['checksum']); extracted=extract_source_rows(pid,f['id'],path); source_rows.extend(extracted)
                if extracted:
                    f['sheets']=sorted({row['sheet'] for row in extracted}); f['readability']='Readable'
        project_finals=[]
        for f in files:
            if f['role']=='Final Quotation': project_finals.extend(extract_final_rows(pid,f['id'],Path(f['path'])))
        final_rows.extend(project_finals)
        alignments.extend(align(name,[r for r in source_rows if r['project_id']==pid],project_finals))

    # Patterns are conservative: one record per observed layout signature; cross-project only when independently recurring.
    signature_projects=defaultdict(set); signature_examples=defaultdict(list)
    for row in source_rows:
        if row['row']==row['formatting'].get('headerRow'):
            sig='|'.join(norm(x) for x in row['cells'] if x not in (None,''))
            if sig: signature_projects[sig].add(row['project_id']); signature_examples[sig].append(row['id'])
    patterns=[]
    for sig,projects in signature_projects.items():
        scope='Cross-Project' if len(projects)>=2 else 'Project-Scoped Experimental'
        patterns.append({'id':stable('histPattern','layout|'+sig),'type':'Recurring Workbook Layout Signature','discipline':'Multi-Discipline','signature':sig,'triggers':{'headerTokens':sig.split('|')},'examples':signature_examples[sig][:10],'behavior':'Map explicit item, description, quantity and unit columns; preserve all rows and blanks.','evidence_count':len(signature_examples[sig]),'confidence':85 if len(projects)>=2 else 60,'scope':scope,'projects':sorted(projects)})

    # Replace only this controlled dataset; no live records are touched.
    with conn:
        for table in ('historical_boq_audit_log','historical_boq_validation_runs','historical_boq_pattern_sources','historical_boq_patterns','historical_boq_decisions','historical_boq_alignments','historical_boq_final_rows','historical_boq_rows','historical_boq_files','historical_boq_projects'):
            conn.execute(f'DELETE FROM {table}')
        for p in project_records:
            conn.execute('INSERT INTO historical_boq_projects (id,name,client,disciplines,project_date,project_status,learning_pair_status,completion_evidence,source_root,review_status) VALUES (?,?,?,?,?,?,?,?,?,?)',(p['id'],p['name'],p['client'],j(p['disciplines']),p['date'],p['status'],p['pair'],j(p['completion']),p['root'],'Needs Review'))
            conn.execute('INSERT INTO historical_boq_audit_log (id,historical_project_id,entity_type,entity_id,action,new_value,reason,actor_user_id) VALUES (?,?,?,?,?,?,?,?)',(stable('histAudit',p['id']),p['id'],'Historical Project',p['id'],'Registered',j(p),'Controlled five-project BOQ learning intake',ACTOR))
        for f in inventory:
            conn.execute('INSERT INTO historical_boq_files (id,historical_project_id,path,file_name,checksum,size_bytes,extension,sheet_names,file_role,source_or_output,revision,role_evidence,role_confidence,human_review_required,readability) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(f['id'],f['project_id'],f['path'],f['name'],f['checksum'],f['size'],f['extension'],j(f['sheets']),f['role'],f['side'],f['revision'],f['evidence'],f['confidence'],f['human_review'],f['readability']))
        for r in source_rows:
            conn.execute('INSERT INTO historical_boq_rows (id,historical_project_id,file_id,sheet_name,row_number,item_number,section_path,original_cells,original_description,original_unit,original_quantity,formulae,discipline,system,location,row_formatting,merged_context,hidden_row,row_type,classification_reason,source_provenance) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(r['id'],r['project_id'],r['file_id'],r['sheet'],r['row'],None if r['item'] is None else str(r['item']),r['section'],j(r['cells']),None if r['description'] is None else str(r['description']),None if r['unit'] is None else str(r['unit']),None if r['quantity'] is None else str(r['quantity']),j(r['formulae']),r['discipline'],r['system'],r['location'],j(r['formatting']),j(r['merged']),int(r['hidden']),r['row_type'],r['reason'],j(r['provenance'])))
        for r in final_rows:
            conn.execute('INSERT INTO historical_boq_final_rows (id,historical_project_id,file_id,page_number,final_row_reference,final_row_type,final_description,final_unit,final_quantity,final_discipline,final_system,final_category,assembly_component,split_merge_decision,exclusion_decision,manufacturer,part_number,accessories,engineer_notes,approval_status,source_provenance) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(r['id'],r['project_id'],r['file_id'],r['page'],r['ref'],r['type'],r['description'],r['unit'],r['quantity'],r['discipline'],r['system'],r['category'],r['assembly'],r['split_merge'],r['exclusion'],r['manufacturer'],r['part_number'],r['accessories'],r['notes'],r['approval'],j(r['provenance'])))
        for a in alignments:
            conn.execute('INSERT INTO historical_boq_alignments (id,historical_project_id,source_row_id,final_row_id,outcome,alignment_method,confidence,evidence,eligible_for_learning) VALUES (?,?,?,?,?,?,?,?,?)',(a['id'],a['project_id'],a['source_row_id'],a['final_row_id'],a['outcome'],a['method'],a['confidence'],j(a['evidence']),a['eligible']))
            if a['eligible']:
                src=next(r for r in source_rows if r['id']==a['source_row_id']); fin=next(r for r in final_rows if r['id']==a['final_row_id'])
                governance={'historicalProjectId':a['project_id'],'alignmentMethod':a['method'],'alignmentConfidence':a['confidence'],'reviewerStatus':'Needs Review','eligibleForProductLearning':False}
                conn.execute('INSERT INTO historical_boq_decisions (id,historical_project_id,alignment_id,source_state,final_state,governance,eligible_for_boq_learning,eligible_for_product_learning,audit_history) VALUES (?,?,?,?,?,?,?,?,?)',(stable('histDecision',a['id']),a['project_id'],a['id'],j(src),j(fin),j(governance),1,0,j([{'action':'Created from trustworthy alignment','actor':ACTOR}])))
        for p in patterns:
            conn.execute('INSERT INTO historical_boq_patterns (id,pattern_type,discipline_scope,layout_signature,trigger_conditions,example_source_rows,expected_behavior,supporting_evidence_count,confidence,scope_status,active_status) VALUES (?,?,?,?,?,?,?,?,?,?,?)',(p['id'],p['type'],p['discipline'],p['signature'],j(p['triggers']),j(p['examples']),p['behavior'],p['evidence_count'],p['confidence'],p['scope'],'Inactive'))
            for project_id in p['projects']:
                conn.execute('INSERT INTO historical_boq_pattern_sources (id,pattern_id,historical_project_id,evidence) VALUES (?,?,?,?)',(stable('histPatternSource',p['id']+'|'+project_id),p['id'],project_id,j({'layoutSignature':p['signature']})))
        for p in project_records:
            eligible=p['pair']=='Complete Learning Pair'
            status='Blocked' if eligible and len([x for x in project_records if x['pair']=='Complete Learning Pair'])<2 else 'Not Applicable'
            blockers=['Reliable holdout metrics require at least two independently grounded complete learning pairs.'] if status=='Blocked' else ['No reliable row-level ground truth.']
            conn.execute('INSERT INTO historical_boq_validation_runs (id,historical_project_id,validation_type,status,metrics,ground_truth_basis,blockers) VALUES (?,?,?,?,?,?,?)',(stable('histValidation',p['id']),p['id'],'Leave-One-Project-Out Holdout',status,j({}),p['pair'],j(blockers)))
    after={t:conn.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0] for t in before}
    if before!=after: raise RuntimeError(f'Live project tables changed: {before} -> {after}')
    summary={'projects':project_records,'inventory':inventory,'sourceRows':source_rows,'finalRows':final_rows,'alignments':alignments,'patterns':patterns,'liveTableCountsBefore':before,'liveTableCountsAfter':after}
    (OUT/'historical_boq_learning.json').write_text(j(summary),encoding='utf-8')
    counts={'projects':len(project_records),'files':len(inventory),'sourceRows':len(source_rows),'finalRows':len(final_rows),'alignments':defaultdict(int),'patterns':len(patterns),'decisions':sum(a['eligible'] for a in alignments)}
    for a in alignments: counts['alignments'][a['outcome']]+=1
    counts['alignments']=dict(counts['alignments'])
    print(j(counts))

if __name__=='__main__': main()
