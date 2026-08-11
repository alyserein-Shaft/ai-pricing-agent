from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs/historical-boq-learning"
DATA_PATH = OUTPUT_DIR / "historical_boq_learning.json"
OUTPUT_PATH = OUTPUT_DIR / "Historical_BOQ_Learning_Review.xlsx"

data = json.loads(DATA_PATH.read_text(encoding="utf-8"))

project_names = {row["id"]: row["name"] for row in data["projects"]}
file_names = {row["id"]: row["name"] for row in data["inventory"]}
source_by_id = {row["id"]: row for row in data["sourceRows"]}
final_by_id = {row["id"]: row for row in data["finalRows"]}

NAVY = "0B2A4A"
HEADER = "DCEAF7"
WHITE = "FFFFFF"
TEXT = "1E2F40"
LINE = "D5E0EC"
LIGHT_LINE = "E7EDF4"

thin = Side(style="thin", color=LIGHT_LINE)
header_border = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)

wb = Workbook()
wb.remove(wb.active)


def safe(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    value = str(value) if not isinstance(value, (int, float, bool)) else value
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def clip(value, size=500):
    return str(value or "")[:size]


def add_sheet(name, headers, rows, widths=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(1, 1, name)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(name="Aptos Display", size=14, bold=True, color=WHITE)
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    for column, header in enumerate(headers, 1):
        cell = ws.cell(3, column, header)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(name="Aptos", size=10, bold=True, color="17324D")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = header_border

    for row_number, row in enumerate(rows, 4):
        for column, value in enumerate(row, 1):
            cell = ws.cell(row_number, column, safe(value))
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(len(rows) + 3, 3)}"

    if widths:
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
    else:
        for index, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(index)].width = min(
                max(len(str(header)) + 3, 12), 32
            )

    return ws


def add_list_validation(ws, cell_range, values):
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
    )
    validation.error = "Select a value from the approved list."
    validation.errorTitle = "Invalid decision"
    ws.add_data_validation(validation)
    validation.add(cell_range)


projects_rows = []
for project in data["projects"]:
    projects_rows.append([
        project["id"],
        project["name"],
        project.get("client", ""),
        ", ".join(project.get("disciplines", [])),
        project.get("date", ""),
        project.get("pairBeforeRecovery", project.get("pair", "")),
        project.get("pair", ""),
        project.get("completion", {}),
        "",
        "",
        "",
        "",
    ])

projects = add_sheet(
    "Projects",
    [
        "Historical Project ID", "Project", "Client", "Disciplines", "Date",
        "Pair Before Recovery", "Current Pair Status", "Evidence",
        "Confirm Pair", "Reviewer Notes", "Reviewer", "Review Date",
    ],
    projects_rows,
    [28, 34, 25, 38, 14, 24, 25, 55, 18, 35, 20, 14],
)
add_list_validation(
    projects,
    f"I4:I{len(projects_rows) + 3}",
    ["Confirm Pair", "Reject Pair", "Needs Evidence"],
)

add_sheet(
    "File Inventory",
    [
        "File ID", "Historical Project ID", "Project", "File Name", "Path",
        "SHA-256", "Size", "Extension", "Role", "Source / Output", "Revision",
        "Evidence Basis", "Confidence", "Human Review", "Readability",
        "Container Path", "Archive Member",
    ],
    [[
        row["id"],
        row["project_id"],
        project_names.get(row["project_id"], ""),
        row["name"],
        row["path"],
        row["checksum"],
        row["size"],
        row["extension"],
        row["role"],
        row["side"],
        row.get("revision", ""),
        row["evidence"],
        row["confidence"],
        "Required" if row.get("human_review") else "Not Required",
        row["readability"],
        row.get("containerPath", ""),
        row.get("archiveMember", ""),
    ] for row in data["inventory"]],
    [28, 28, 32, 38, 60, 42, 14, 10, 24, 18, 12, 55, 12, 15, 26, 60, 45],
)

add_sheet(
    "Source-Final Pairs",
    [
        "Historical Project ID", "Project", "Authoritative Source Basis",
        "Revised Sources", "Reviewed Output Basis", "Final Cost Sheet",
        "Final Quotation", "Pair Classification", "Evidence",
        "Reviewer Decision", "Reviewer Notes", "Reviewer", "Review Date",
    ],
    [[
        project["id"],
        project["name"],
        "See File Inventory",
        "See File Inventory",
        "See File Inventory",
        "Present only where readable/confirmed",
        "Issued PDF present where registered",
        project.get("pair", ""),
        project.get("completion", {}).get(
            "pairEvidence",
            "Source and reviewed output require human confirmation.",
        ),
        "",
        "",
        "",
        "",
    ] for project in data["projects"]],
    [28, 34, 40, 32, 40, 28, 25, 26, 60, 20, 35, 20, 14],
)

add_sheet(
    "Approved Ground Truth",
    [
        "Alignment ID", "Historical Project ID", "Project", "Source Row ID",
        "Final Row ID", "Approval", "Reviewer", "Review Date", "Audit Reference",
    ],
    [[
        row["alignmentId"],
        row["projectId"],
        project_names.get(row["projectId"], ""),
        row["sourceRowId"],
        row["finalRowId"],
        "Approved Ground Truth",
        row.get("reviewer", ""),
        row.get("reviewDate", ""),
        row.get("auditId", ""),
    ] for row in data.get("approvedGroundTruth", [])],
    [28, 28, 32, 28, 28, 22, 20, 14, 28],
)

review_headers = [
    "Alignment ID", "Historical Project ID", "Project", "Source Row ID",
    "Source File", "Sheet", "Row", "Item", "Source Description",
    "Source Unit", "Source Quantity", "Final Row ID", "Final File", "Page",
    "Final Description", "Final Unit", "Final Quantity", "Original Outcome",
    "System Suggestion", "Score", "Unit Difference", "Quantity Difference",
    "Conflicting Signals", "Reviewer Decision", "Reviewer Notes",
    "Reviewer", "Review Date",
]


def review_row(review):
    source = source_by_id.get(review["sourceRowId"], {})
    final = final_by_id.get(review["finalRowId"], {})
    score = review.get("scoreComponents", {})
    return [
        review["alignmentId"],
        review["projectId"],
        project_names.get(review["projectId"], ""),
        review["sourceRowId"],
        file_names.get(source.get("file_id"), ""),
        source.get("sheet", ""),
        source.get("row", ""),
        source.get("item", ""),
        clip(source.get("description")),
        source.get("unit", ""),
        source.get("quantity", ""),
        review["finalRowId"],
        file_names.get(final.get("file_id"), ""),
        final.get("page", ""),
        clip(final.get("description")),
        final.get("unit", ""),
        final.get("quantity", ""),
        review.get("originalOutcome", ""),
        review.get("reviewSuggestion", ""),
        score.get("score", ""),
        score.get("unitDifference", ""),
        score.get("quantityDifference", ""),
        "; ".join(review.get("conflictingSignals", [])),
        review.get("humanDecision", ""),
        review.get("reviewerNotes", ""),
        review.get("reviewer", ""),
        review.get("reviewDate", ""),
    ]


alignment_reviews = data.get("alignmentReviews", [])

exact_strong_rows = [
    review_row(row)
    for row in alignment_reviews
    if row.get("originalOutcome") in ("Exact Alignment", "Strong Alignment")
]

exact_strong = add_sheet(
    "Exact and Strong Review",
    review_headers,
    exact_strong_rows,
    [28, 28, 32, 28, 35, 16, 10, 12, 55, 12, 14, 28, 38, 10, 55, 12, 14, 18, 20, 10, 24, 24, 45, 24, 35, 20, 14],
)
add_list_validation(
    exact_strong,
    f"X4:X{max(len(exact_strong_rows) + 3, 4)}",
    [
        "Approved Ground Truth",
        "Needs Engineer Review",
        "Rejected Alignment",
        "Duplicate Alignment",
        "Possible Split",
        "Possible Merge",
    ],
)

possible_rows = [
    review_row(row)
    for row in alignment_reviews
    if row.get("originalOutcome") == "Possible Alignment"
]

possible = add_sheet(
    "Possible Alignments",
    review_headers,
    possible_rows,
    [28, 28, 32, 28, 35, 16, 10, 12, 55, 12, 14, 28, 38, 10, 55, 12, 14, 18, 20, 10, 24, 24, 45, 24, 35, 20, 14],
)
add_list_validation(
    possible,
    f"X4:X{max(len(possible_rows) + 3, 4)}",
    [
        "Needs Engineer Review",
        "Rejected Alignment",
        "Duplicate Alignment",
        "Possible Split",
        "Possible Merge",
    ],
)

candidate_headers = []
for index in range(1, 6):
    candidate_headers += [
        f"Candidate {index} Final Row ID",
        f"Candidate {index} Score",
        f"Candidate {index} Basis",
        f"Candidate {index} Unit Difference",
        f"Candidate {index} Quantity Difference",
    ]

unresolved_rows = []
for row in data.get("unresolvedAssistance", []):
    candidates = []
    for index in range(5):
        candidate = row.get("candidates", [])[index] if index < len(row.get("candidates", [])) else None
        if candidate:
            candidates += [
                candidate.get("finalRowId", ""),
                candidate.get("score", ""),
                {
                    "label": candidate.get("label"),
                    "descriptionSimilarity": candidate.get("descriptionSimilarity"),
                    "itemEqual": candidate.get("itemEqual"),
                    "unitEqual": candidate.get("unitEqual"),
                    "quantityEqual": candidate.get("quantityEqual"),
                    "sharedTechnicalTerms": candidate.get("sharedTechnicalTerms"),
                },
                candidate.get("unitDifference", ""),
                candidate.get("quantityDifference", ""),
            ]
        else:
            candidates += ["", "", "", "", ""]

    unresolved_rows.append([
        row["sourceRowId"],
        row["projectId"],
        project_names.get(row["projectId"], ""),
        file_names.get(row.get("fileId"), ""),
        row.get("sheet", ""),
        row.get("row", ""),
        row.get("section", ""),
        row.get("sourceDescription", ""),
        row.get("sourceUnit", ""),
        row.get("sourceQuantity", ""),
        row.get("assistanceOutcome", ""),
        *candidates,
        row.get("reviewerSelectedCandidate", ""),
        row.get("alignmentDecision", ""),
        row.get("reviewerNotes", ""),
        row.get("reviewer", ""),
        row.get("reviewDate", ""),
    ])

unresolved_headers = [
    "Stable Source Row ID", "Historical Project ID", "Project", "Source File",
    "Sheet", "Row", "Section", "Source Description", "Source Unit",
    "Source Quantity", "Assistance Outcome",
    *candidate_headers,
    "Reviewer Selected Candidate", "Alignment Decision", "Reviewer Notes",
    "Reviewer", "Review Date",
]

unresolved = add_sheet(
    "Unresolved Review Queue",
    unresolved_headers,
    unresolved_rows,
    [28 if index < 3 else 55 if index == 7 else 38 if index == 3 else 16 for index in range(len(unresolved_headers))],
)

decision_column = unresolved_headers.index("Alignment Decision") + 1
decision_letter = get_column_letter(decision_column)
add_list_validation(
    unresolved,
    f"{decision_letter}4:{decision_letter}{max(len(unresolved_rows) + 3, 4)}",
    [
        "Likely Exact", "Likely Strong", "Possible", "No Candidate",
        "Possible Split", "Possible Merge", "Likely Excluded",
    ],
)

add_sheet(
    "Split-Merge Candidates",
    review_headers,
    [review_row(row) for row in data.get("splitMergeCandidates", [])],
    [28, 28, 32, 28, 35, 16, 10, 12, 55, 12, 14, 28, 38, 10, 55, 12, 14, 18, 20, 10, 24, 24, 45, 24, 35, 20, 14],
)

add_sheet(
    "Exclusions",
    [
        "Final Row ID", "Historical Project ID", "Project", "File", "Page",
        "Description", "System Indicator", "Reviewer Decision",
        "Reviewer Notes", "Reviewer", "Review Date",
    ],
    [[
        row["id"],
        row["project_id"],
        project_names.get(row["project_id"], ""),
        file_names.get(row["file_id"], ""),
        row.get("page", ""),
        clip(row.get("description")),
        row.get("exclusion", ""),
        "",
        "",
        "",
        "",
    ] for row in data["finalRows"] if row.get("exclusion")],
    [28, 28, 32, 38, 10, 55, 20, 22, 35, 20, 14],
)

add_sheet(
    "Extracted Patterns",
    [
        "Pattern ID", "Pattern Type", "Scope", "Discipline", "Signature",
        "Trigger Conditions", "Expected Behavior", "Evidence Count",
        "Confidence", "Source Projects", "Active Status",
    ],
    [[
        row["id"],
        row["type"],
        row["scope"],
        row["discipline"],
        row["signature"],
        row["triggers"],
        row["behavior"],
        row["evidence_count"],
        row["confidence"],
        ", ".join(project_names.get(project_id, project_id) for project_id in row["projects"]),
        "Inactive — no approved ground truth",
    ] for row in data["patterns"]],
    [28, 32, 28, 20, 55, 45, 55, 16, 12, 45, 30],
)

pattern_review_rows = [[
    row["id"], row["type"], row["scope"], "", "", "", "", ""
] for row in data["patterns"]]

pattern_review = add_sheet(
    "Pattern Review",
    [
        "Pattern ID", "Pattern Type", "Current Scope", "Reviewer Decision",
        "Restriction", "Reviewer Notes", "Reviewer", "Review Date",
    ],
    pattern_review_rows,
    [28, 32, 28, 22, 28, 40, 20, 14],
)
add_list_validation(
    pattern_review,
    f"D4:D{max(len(pattern_review_rows) + 3, 4)}",
    [
        "Keep Inactive",
        "Approve after Ground Truth",
        "Reject Pattern",
        "Needs Evidence",
    ],
)

add_sheet(
    "Blocked Files",
    [
        "Project", "File Path", "Blocker", "Readable Alternate Evidence",
        "Human Action Required",
    ],
    [[
        row.get("project", ""),
        row.get("path", ""),
        row.get("reason", ""),
        row.get("alternateEvidence", ""),
        "Provide password or a readable unencrypted cost-sheet export.",
    ] for row in data.get("blockedFiles", [])],
    [32, 65, 52, 48, 48],
)

add_sheet(
    "Instructions",
    ["Step", "Instruction", "Safety Boundary"],
    [
        [1, "Confirm each source/final pair using the stable historical project ID.", "Filenames and folder placement are evidence, not final authority."],
        [2, "Review all Exact and Strong rows. Current candidates may reuse the same final row.", "No score or system suggestion grants approval."],
        [3, "Review Possible Alignments and record a substantive reason.", "Possible rows cannot be promoted without explicit evidence."],
        [4, "Use the Unresolved Review Queue and select candidates only within the same project.", "Candidate ranking never auto-approves."],
        [5, "Enter reviewer identity, notes and review date for every decision.", "Only explicit Approved Ground Truth becomes learning evidence."],
        [6, "Keep patterns inactive until rebuilt from approved ground truth.", "Single-project evidence remains project-scoped."],
        [7, "Run holdout only after three complete grounded projects exist.", "Do not simulate missing ground truth or process the live project."],
    ],
    [10, 85, 80],
)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT_PATH)

print(json.dumps({
    "workbook": str(OUTPUT_PATH),
    "size": OUTPUT_PATH.stat().st_size,
    "sheets": wb.sheetnames,
    "alignmentReviews": len(alignment_reviews),
    "unresolvedRows": len(unresolved_rows),
    "blockedFiles": len(data.get("blockedFiles", [])),
}, ensure_ascii=False))
