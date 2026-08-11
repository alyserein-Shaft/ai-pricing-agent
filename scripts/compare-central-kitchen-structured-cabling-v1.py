from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
CORNING_RFQ = BASE / "data-networking-engineer-rfq.xlsx"
FINAL_QUOTATION = BASE / "final-quotation.pdf"

OUTPUT_DIR = Path("outputs/central-kitchen-structured-cabling")
OUTPUT_FILE = (
    OUTPUT_DIR
    / "Central_Kitchen_Structured_Cabling_Comparison_v1.xlsx"
)


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()

    replacements = {
        "cat.6a": "cat6a",
        "cat 6a": "cat6a",
        "cat.6": "cat6",
        "cat 6": "cat6",
        "rj-45": "rj45",
        "rj 45": "rj45",
        "fibre": "fiber",
        "patchpanel": "patch panel",
        "face plate": "faceplate",
        "double gang": "double",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)

    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def longest_text(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if (
            isinstance(value, str)
            and len(clean(value)) >= 5
            and not clean(value).startswith("=")
        )
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    numbers: list[float] = []

    for value in values:
        number = numeric(value)

        if number is not None:
            numbers.append(number)

    if not numbers:
        return None

    for number in reversed(numbers):
        if number >= 1:
            return number

    return numbers[-1]


def find_unit(values: list[Any]) -> str:
    units = {
        "no",
        "nos",
        "no.",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in units:
            return clean(value)

    return ""


def find_item_number(values: list[Any]) -> str:
    for value in values[:6]:
        text = clean(value)

        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
            return text

    return ""


def classify_source_item(description: str) -> tuple[str, int]:
    text = normalize(description)

    if "15u rack" in text:
        return "15U Rack", 0

    if "18u rack" in text:
        return "18U Rack", 0

    if "24u rack" in text and "midf" in text:
        return "24U MIDF Rack", 0

    if "24u rack" in text:
        return "24U IDF Rack", 0

    if "wireless data outlet" in text and "ceiling" in text:
        return "Ceiling Wireless Data Outlet", 1

    if "wall mounted wireless data outlet" in text:
        return "Wall Wireless Data Outlet", 1

    if "telephone and data" in text and "2 no" in text:
        return "Telephone + Data Double Outlet", 2

    if "double" in text and "data outlet" in text:
        return "Double Data Outlet", 2

    if "data outlet" in text and "rj45" in text:
        return "Single Data Outlet", 1

    if "hdmi outlet" in text:
        return "HDMI Outlet", 0

    return "", 0


def extract_source_boq() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    ws = workbook["Div. 27"]

    records: list[dict[str, Any]] = []

    for row_number in range(10, 124):
        values = row_values(ws, row_number)
        description = longest_text(values)
        quantity = find_quantity(values)

        if not description or quantity is None:
            continue

        category, ports_per_item = classify_source_item(description)

        if not category:
            continue

        estimated_ports = (
            float(quantity) * ports_per_item
            if ports_per_item > 0
            else 0
        )

        records.append(
            {
                "source_category": category,
                "source_row": row_number,
                "item_number": find_item_number(values),
                "description": description,
                "quantity": quantity,
                "unit": find_unit(values),
                "ports_per_item": ports_per_item,
                "estimated_rj45_ports": estimated_ports,
            }
        )

    return records


def classify_rfq_item(description: str, part_number: str) -> str:
    text = normalize(f"{part_number} {description}")

    if "faceplate" in text and "1 jack" in text:
        return "1-Port Faceplate"

    if "faceplate" in text and "2 jack" in text:
        return "2-Port Faceplate"

    if "cat6a" in text and "jack" in text and "patch panel" in text:
        return "Patch Panel Jack"

    if "cat6a" in text and "jack" in text:
        return "Outlet Jack"

    if "male module connector" in text:
        return "Field RJ45 Connector"

    if "patch panel 24p" in text:
        return "24-Port Patch Panel"

    if "patch panel 48p" in text:
        return "48-Port Patch Panel"

    if "cable management" in text:
        return "Cable Manager"

    if "patchcord" in text and "cat6a" in text:
        return "Copper Patch Cord"

    if "fiber patch panel" in text:
        return "Fiber Patch Panel"

    if "fo patchcord" in text or (
        "om4" in text and "lc lc" in text
    ):
        return "Fiber Patch Cord"

    if "hdmi outlet" in text:
        return "HDMI Outlet"

    return "Other Passive Component"


def extract_corning_rfq() -> list[dict[str, Any]]:
    workbook = load_workbook(CORNING_RFQ, data_only=False)
    ws = workbook["Sheet1"]

    records: list[dict[str, Any]] = []

    for row_number in range(2, ws.max_row + 1):
        part_number = clean(ws.cell(row=row_number, column=1).value)
        description = clean(ws.cell(row=row_number, column=2).value)
        quantity = numeric(ws.cell(row=row_number, column=3).value)

        if not description or quantity is None:
            continue

        category = classify_rfq_item(description, part_number)

        capacity = 0

        if category == "1-Port Faceplate":
            capacity = int(quantity)

        elif category == "2-Port Faceplate":
            capacity = int(quantity) * 2

        elif category == "24-Port Patch Panel":
            capacity = int(quantity) * 24

        elif category == "48-Port Patch Panel":
            capacity = int(quantity) * 48

        records.append(
            {
                "rfq_row": row_number,
                "part_number": part_number,
                "description": description,
                "quantity": quantity,
                "rfq_category": category,
                "calculated_port_capacity": capacity,
            }
        )

    return records


def extract_pdf_lines() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    lines: list[dict[str, Any]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            line = clean(raw_line)

            if line:
                lines.append(
                    {
                        "page": page_number,
                        "line_number": line_number,
                        "text": line,
                    }
                )

    return lines


def extract_final_passive_lines() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    keywords = (
        "corning",
        "cat6",
        "cat 6",
        "rj45",
        "rj-45",
        "patch panel",
        "patch cord",
        "faceplate",
        "fiber patch",
        "fibre patch",
        "cable management",
        "cable manager",
        "rack",
        "pdu",
        "om4",
        "lc-lc",
        "lc lc",
    )

    for entry in extract_pdf_lines():
        normalized = normalize(entry["text"])

        if not any(normalize(keyword) in normalized for keyword in keywords):
            continue

        records.append(
            {
                "page": entry["page"],
                "line_number": entry["line_number"],
                "description": entry["text"],
                "detected_category": classify_rfq_item(
                    entry["text"],
                    "",
                ),
            }
        )

    return records


def sum_source(
    records: list[dict[str, Any]],
    categories: set[str],
    field: str = "quantity",
) -> float:
    return sum(
        float(record[field])
        for record in records
        if record["source_category"] in categories
    )


def sum_rfq(
    records: list[dict[str, Any]],
    category: str,
    field: str = "quantity",
) -> float:
    return sum(
        float(record[field])
        for record in records
        if record["rfq_category"] == category
    )


def build_reconciliation(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_racks = sum_source(
        source_records,
        {
            "15U Rack",
            "18U Rack",
            "24U MIDF Rack",
            "24U IDF Rack",
        },
    )

    source_ports = sum(
        float(record["estimated_rj45_ports"])
        for record in source_records
    )

    rfq_one_port_faceplates = sum_rfq(
        rfq_records,
        "1-Port Faceplate",
    )

    rfq_two_port_faceplates = sum_rfq(
        rfq_records,
        "2-Port Faceplate",
    )

    rfq_faceplate_capacity = (
        rfq_one_port_faceplates
        + (rfq_two_port_faceplates * 2)
    )

    outlet_jacks = sum_rfq(
        rfq_records,
        "Outlet Jack",
    )

    patch_panel_jacks = sum_rfq(
        rfq_records,
        "Patch Panel Jack",
    )

    panel_24_capacity = sum_rfq(
        rfq_records,
        "24-Port Patch Panel",
        "calculated_port_capacity",
    )

    panel_48_capacity = sum_rfq(
        rfq_records,
        "48-Port Patch Panel",
        "calculated_port_capacity",
    )

    return [
        {
            "metric": "Source rack quantity",
            "source_value": source_racks,
            "rfq_value": "",
            "difference": "",
            "status": "Reference",
            "observation": "Source contains 11 racks.",
        },
        {
            "metric": "Fiber patch panel quantity",
            "source_value": source_racks,
            "rfq_value": sum_rfq(
                rfq_records,
                "Fiber Patch Panel",
            ),
            "difference": (
                sum_rfq(rfq_records, "Fiber Patch Panel")
                - source_racks
            ),
            "status": "PASS"
            if sum_rfq(rfq_records, "Fiber Patch Panel") == source_racks
            else "REVIEW",
            "observation": "Observed one fiber patch panel per rack.",
        },
        {
            "metric": "Source estimated RJ45 ports",
            "source_value": source_ports,
            "rfq_value": "",
            "difference": "",
            "status": "Estimated",
            "observation": (
                "Calculated from outlet type; requires specs/drawing review."
            ),
        },
        {
            "metric": "RFQ faceplate port capacity",
            "source_value": source_ports,
            "rfq_value": rfq_faceplate_capacity,
            "difference": rfq_faceplate_capacity - source_ports,
            "status": "REVIEW",
            "observation": "Faceplate capacity versus BOQ port estimate.",
        },
        {
            "metric": "RFQ outlet-side jacks",
            "source_value": source_ports,
            "rfq_value": outlet_jacks,
            "difference": outlet_jacks - source_ports,
            "status": "REVIEW",
            "observation": (
                "Does not include jacks explicitly marked for patch panels."
            ),
        },
        {
            "metric": "24-port patch-panel capacity",
            "source_value": "",
            "rfq_value": panel_24_capacity,
            "difference": "",
            "status": "Reference",
            "observation": "27 panels × 24 ports.",
        },
        {
            "metric": "48-port patch-panel capacity",
            "source_value": "",
            "rfq_value": panel_48_capacity,
            "difference": "",
            "status": "Reference",
            "observation": "4 panels × 48 ports.",
        },
        {
            "metric": "Total patch-panel capacity",
            "source_value": source_ports,
            "rfq_value": panel_24_capacity + panel_48_capacity,
            "difference": (
                panel_24_capacity
                + panel_48_capacity
                - source_ports
            ),
            "status": "REVIEW",
            "observation": (
                "Capacity likely includes spare ports or additional ELV systems."
            ),
        },
        {
            "metric": "Patch-panel jacks",
            "source_value": panel_24_capacity + panel_48_capacity,
            "rfq_value": patch_panel_jacks,
            "difference": (
                patch_panel_jacks
                - panel_24_capacity
                - panel_48_capacity
            ),
            "status": "PASS"
            if patch_panel_jacks
            == panel_24_capacity + panel_48_capacity
            else "REVIEW",
            "observation": "Jacks versus installed patch-panel capacity.",
        },
        {
            "metric": "Copper patch cords",
            "source_value": panel_24_capacity + panel_48_capacity,
            "rfq_value": sum_rfq(
                rfq_records,
                "Copper Patch Cord",
            ),
            "difference": (
                sum_rfq(rfq_records, "Copper Patch Cord")
                - panel_24_capacity
                - panel_48_capacity
            ),
            "status": "PASS"
            if sum_rfq(rfq_records, "Copper Patch Cord")
            == panel_24_capacity + panel_48_capacity
            else "REVIEW",
            "observation": "One patch cord per populated patch-panel port.",
        },
    ]


def build_validation(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_racks = sum_source(
        source_records,
        {
            "15U Rack",
            "18U Rack",
            "24U MIDF Rack",
            "24U IDF Rack",
        },
    )

    source_ports = sum(
        float(record["estimated_rj45_ports"])
        for record in source_records
    )

    panel_24_capacity = sum_rfq(
        rfq_records,
        "24-Port Patch Panel",
        "calculated_port_capacity",
    )

    panel_48_capacity = sum_rfq(
        rfq_records,
        "48-Port Patch Panel",
        "calculated_port_capacity",
    )

    checks = [
        {
            "check": "Source measurable items",
            "expected": 10,
            "actual": len(source_records),
        },
        {
            "check": "Source rack quantity",
            "expected": 11,
            "actual": source_racks,
        },
        {
            "check": "Source estimated RJ45 ports",
            "expected": 242,
            "actual": source_ports,
        },
        {
            "check": "RFQ fiber patch panels",
            "expected": 11,
            "actual": sum_rfq(
                rfq_records,
                "Fiber Patch Panel",
            ),
        },
        {
            "check": "RFQ 24-port panel capacity",
            "expected": 648,
            "actual": panel_24_capacity,
        },
        {
            "check": "RFQ 48-port panel capacity",
            "expected": 192,
            "actual": panel_48_capacity,
        },
        {
            "check": "RFQ total panel capacity",
            "expected": 840,
            "actual": panel_24_capacity + panel_48_capacity,
        },
        {
            "check": "RFQ patch-panel jacks",
            "expected": 840,
            "actual": sum_rfq(
                rfq_records,
                "Patch Panel Jack",
            ),
        },
        {
            "check": "RFQ copper patch cords",
            "expected": 840,
            "actual": sum_rfq(
                rfq_records,
                "Copper Patch Cord",
            ),
        },
    ]

    for check in checks:
        check["status"] = (
            "PASS"
            if float(check["expected"]) == float(check["actual"])
            else "FAIL"
        )

    return checks


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append([record.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        maximum = max(
            len(clean(cell.value))
            for cell in column_cells
        )

        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = min(
            max(maximum + 2, 12),
            55,
        )


def main() -> None:
    for path in (SOURCE_BOQ, CORNING_RFQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_boq()
    rfq_records = extract_corning_rfq()
    final_lines = extract_final_passive_lines()

    reconciliation = build_reconciliation(
        source_records,
        rfq_records,
    )

    validation = build_validation(
        source_records,
        rfq_records,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Source Structured Cabling"),
        [
            "source_category",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "ports_per_item",
            "estimated_rj45_ports",
        ],
        source_records,
    )

    write_sheet(
        workbook.create_sheet("Corning RFQ Components"),
        [
            "rfq_row",
            "part_number",
            "description",
            "quantity",
            "rfq_category",
            "calculated_port_capacity",
        ],
        rfq_records,
    )

    write_sheet(
        workbook.create_sheet("Final Passive Evidence"),
        [
            "page",
            "line_number",
            "description",
            "detected_category",
        ],
        final_lines,
    )

    write_sheet(
        workbook.create_sheet("Quantity Reconciliation"),
        [
            "metric",
            "source_value",
            "rfq_value",
            "difference",
            "status",
            "observation",
        ],
        reconciliation,
    )

    write_sheet(
        workbook.create_sheet("Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        validation,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source measurable items: {len(source_records)}")
    print(f"Corning RFQ components: {len(rfq_records)}")
    print(f"Final passive evidence lines: {len(final_lines)}")
    print()

    for check in validation:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} '
            f'actual={check["actual"]}'
        )

    print()
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
