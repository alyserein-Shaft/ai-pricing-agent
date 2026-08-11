from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
ENGINEER_RFQ = BASE / "cctv-engineer-rfq.xlsx"
FINAL_QUOTATION = (
    BASE
    / "final-quotation"
    / "Q1067-626-LCU- Central Kitchen - Makkah.pdf"
)

OUTPUT_DIR = Path("outputs/central-kitchen-cctv-comparison")
OUTPUT_FILE = OUTPUT_DIR / "Central_Kitchen_CCTV_Comparison_v3.xlsx"


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()

    replacements = {
        "anti-fog": "anti fog",
        "antifog": "anti fog",
        "anti  fog": "anti fog",
        "dom type": "dome type",
        "dom camera": "dome camera",
        "n.v.r": "nvr",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)

    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def longest_text(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if isinstance(value, str) and len(clean(value)) >= 8
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    candidates: list[tuple[int, float]] = []

    for index, value in enumerate(values):
        number = numeric(value)

        if number is not None:
            candidates.append((index, number))

    if not candidates:
        return None

    for _, number in reversed(candidates):
        if number >= 1:
            return number

    return candidates[-1][1]


def find_unit(values: list[Any]) -> str:
    units = {
        "no",
        "nos",
        "no.",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in units:
            return clean(value)

    return ""


def find_item_number(values: list[Any]) -> str:
    for value in values[:6]:
        text = clean(value)

        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
            return text

    return ""


def classify_camera_description(description: str) -> str:
    text = normalize(description)

    has_indoor = "indoor" in text
    has_outdoor = "outdoor" in text
    has_dome = "dome" in text or re.search(r"\bdom\b", text) is not None
    has_bullet = "bullet" in text
    has_anti_fog = "anti fog" in text or "defog" in text
    has_pole = "pole" in text
    has_nvr = (
        "network video recorder" in text
        or re.search(r"\bnvr\b", text) is not None
    )

    # Specific rules must run before general rules.
    if has_anti_fog and has_bullet and has_pole:
        return "Anti-fog Bullet Camera with Pole"

    if has_anti_fog and has_bullet:
        return "Anti-fog Bullet Camera"

    if has_anti_fog and has_dome:
        return "Anti-fog Dome Camera"

    # Some source descriptions say anti-fog camera without explicitly
    # repeating "bullet"; wall-mounted is treated as bullet.
    if has_anti_fog and "wall mounted" in text and has_pole:
        return "Anti-fog Bullet Camera with Pole"

    if has_anti_fog and "wall mounted" in text:
        return "Anti-fog Bullet Camera"

    if has_anti_fog and "ceiling mounted" in text:
        return "Anti-fog Dome Camera"

    if has_indoor and has_dome:
        return "Indoor Dome Camera"

    if has_outdoor and has_dome:
        return "Outdoor Dome Camera"

    if has_outdoor and has_bullet:
        return "Outdoor Bullet Camera"

    if has_nvr:
        return "NVR"

    return ""


def extract_source_cctv() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if "28" not in sheet_name and "cctv" not in sheet_name.lower():
            continue

        ws = workbook[sheet_name]
        inside_cctv = False

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))
            normalized = normalize(joined)

            if not joined:
                continue

            if any(
                phrase in normalized
                for phrase in (
                    "closed circuit television",
                    "cctv system",
                    "video surveillance",
                )
            ):
                inside_cctv = True

            if inside_cctv and any(
                phrase in normalized
                for phrase in (
                    "card access system",
                    "access control system",
                    "fire alarm system",
                    "public address system",
                    "data networking",
                    "structured cabling",
                )
            ):
                if "closed circuit television" not in normalized:
                    break

            if not inside_cctv:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)

            if not description or quantity is None:
                continue

            source_item = classify_camera_description(description)

            if not source_item:
                continue

            records.append(
                {
                    "source_item": source_item,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "item_number": find_item_number(values),
                    "description": description,
                    "quantity": quantity,
                    "unit": find_unit(values),
                    "raw_row": joined,
                }
            )

    return records


def extract_engineer_rfq() -> list[dict[str, Any]]:
    workbook = load_workbook(ENGINEER_RFQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))

            if not joined:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)

            if not description or quantity is None:
                continue

            mapped_source_item = classify_camera_description(description)
            normalized = normalize(description)

            if not mapped_source_item:
                if "workstation" in normalized:
                    mapped_source_item = "Workstation"
                elif "monitor" in normalized:
                    mapped_source_item = "Monitor"
                elif "server" in normalized:
                    mapped_source_item = "Server"
                else:
                    continue

            part_number = ""

            for value in values:
                text = clean(value)

                if (
                    len(text) >= 4
                    and re.search(r"[A-Za-z]", text)
                    and re.search(r"\d", text)
                    and " " not in text
                    and not text.startswith("=")
                ):
                    part_number = text
                    break

            records.append(
                {
                    "mapped_source_item": mapped_source_item,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "part_number": part_number,
                    "description": description,
                    "quantity": quantity,
                    "unit": find_unit(values),
                    "raw_row": joined,
                }
            )

    return records


def extract_pdf_lines() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    lines: list[dict[str, Any]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for line_number, raw_line in enumerate(text.splitlines(), start=1):
            line = clean(raw_line)

            if not line:
                continue

            lines.append(
                {
                    "page": page_number,
                    "line_number": line_number,
                    "text": line,
                }
            )

    return lines


def looks_like_price_line(text: str) -> bool:
    return bool(
        re.search(
            r"(?:^|\s)"
            r"(\d+(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
            r"(?:\s|$)",
            text,
        )
    )


def parse_price_line(text: str) -> tuple[float | None, float | None, float | None]:
    matches = list(
        re.finditer(
            r"(?:^|\s)"
            r"(\d+(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+"
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
            r"(?:\s|$)",
            text,
        )
    )

    if not matches:
        return None, None, None

    match = matches[-1]

    return (
        float(match.group(1)),
        float(match.group(2).replace(",", "")),
        float(match.group(3).replace(",", "")),
    )


def extract_part_number(text: str) -> str:
    compact = clean(text)

    candidates = re.findall(
        r"\b[A-Z0-9]+(?:[-/][A-Z0-9().]+)+\b",
        compact,
        flags=re.IGNORECASE,
    )

    rejected = {
        "30-days",
        "90-days",
        "256-ch",
        "6-mp",
        "10-tb",
    }

    for candidate in candidates:
        normalized = candidate.lower()

        if normalized in rejected:
            continue

        if re.search(r"[A-Za-z]", candidate) and re.search(r"\d", candidate):
            return candidate

    return ""


def classify_final_item(description: str) -> str:
    text = normalize(description)

    if "junction box" in text:
        return "Junction Box"

    if "pole mount" in text or "pole mounting" in text:
        return "Pole Mount"

    if "hard disk" in text or re.search(r"\bhdd\b", text):
        return "Storage HDD"

    if "defog" in text and "bullet" in text:
        return "Anti-fog Bullet Camera"

    if "anti fog" in text and "bullet" in text:
        return "Anti-fog Bullet Camera"

    if "defog" in text and "dome" in text:
        return "Anti-fog Dome Camera"

    if "anti fog" in text and "dome" in text:
        return "Anti-fog Dome Camera"

    if "bullet network camera" in text or "bullet camera" in text:
        return "Outdoor Bullet Camera"

    if "dome network camera" in text or "dome camera" in text:
        return "Dome Camera Consolidated"

    if "network video recorder" in text or re.search(r"\bnvr\b", text):
        return "NVR"

    if "workstation" in text:
        return "Workstation"

    if "monitor" in text:
        return "Monitor"

    if "windows server" in text and "license" in text:
        return "Windows Server License"

    if "server" in text and "license" not in text:
        return "Server"

    if "channel license" in text or "channel licence" in text:
        return "VMS Channel License"

    if "license" in text or "licence" in text:
        return "VMS Base License"

    if (
        "testing and commissioning" in text
        or "testing & commissioning" in text
        or "configuration and commissioning" in text
    ):
        return "Testing & Commissioning"

    return ""


def extract_final_cctv() -> list[dict[str, Any]]:
    lines = extract_pdf_lines()
    records: list[dict[str, Any]] = []

    inside_cctv = False
    buffer: list[str] = []
    buffer_page = 0

    start_markers = (
        "closed circuit television",
        "cctv system",
        "video surveillance",
    )

    stop_markers = (
        "access control system",
        "public address system",
        "fire alarm system",
        "data networking",
        "structured cabling",
    )

    def flush() -> None:
        nonlocal buffer, buffer_page

        if not buffer:
            return

        full_text = clean(" ".join(buffer))
        category = classify_final_item(full_text)

        if not category:
            buffer = []
            buffer_page = 0
            return

        quantity, unit_price, item_price = parse_price_line(full_text)

        records.append(
            {
                "page": buffer_page,
                "mapped_category": category,
                "part_number": extract_part_number(full_text),
                "description": full_text,
                "quantity": quantity,
                "unit_price": unit_price,
                "item_price": item_price,
            }
        )

        buffer = []
        buffer_page = 0

    for entry in lines:
        line = entry["text"]
        normalized = normalize(line)

        if any(marker in normalized for marker in start_markers):
            inside_cctv = True
            continue

        if inside_cctv and any(marker in normalized for marker in stop_markers):
            flush()
            inside_cctv = False
            continue

        if not inside_cctv:
            continue

        # Totals and commercial text end a pending item.
        if any(
            marker in normalized
            for marker in (
                "cctv total",
                "section total",
                "payment terms",
                "payment condition",
                "delivery period",
                "warranty",
            )
        ):
            flush()
            continue

        if not buffer:
            buffer = [line]
            buffer_page = entry["page"]
        else:
            buffer.append(line)

        # Product records are considered complete only when the line
        # contains quantity, unit price and total price.
        if looks_like_price_line(line):
            flush()

    flush()

    return records


def build_relationships(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    source_by_name = {
        record["source_item"]: record
        for record in source_records
    }

    rfq_by_name: dict[str, list[dict[str, Any]]] = {}
    final_by_category: dict[str, list[dict[str, Any]]] = {}

    for record in rfq_records:
        rfq_by_name.setdefault(
            record["mapped_source_item"],
            [],
        ).append(record)

    for record in final_records:
        final_by_category.setdefault(
            record["mapped_category"],
            [],
        ).append(record)

    relationships: list[dict[str, Any]] = []

    for source_name, source in source_by_name.items():
        rfq_matches = rfq_by_name.get(source_name, [])

        if source_name in {"Indoor Dome Camera", "Outdoor Dome Camera"}:
            final_categories = [
                "Dome Camera Consolidated",
                "Junction Box",
            ]
            relationship_type = "Technical Consolidation"

        elif source_name == "Outdoor Bullet Camera":
            final_categories = [
                "Outdoor Bullet Camera",
                "Junction Box",
            ]
            relationship_type = "Product + Accessory"

        elif source_name == "Anti-fog Bullet Camera":
            final_categories = [
                "Anti-fog Bullet Camera",
            ]
            relationship_type = "Technical Consolidation"

        elif source_name == "Anti-fog Bullet Camera with Pole":
            final_categories = [
                "Anti-fog Bullet Camera",
                "Pole Mount",
            ]
            relationship_type = "Product + Pole Accessory"

        elif source_name == "Anti-fog Dome Camera":
            final_categories = [
                "Anti-fog Dome Camera",
                "Junction Box",
            ]
            relationship_type = "Product + Accessory"

        elif source_name == "NVR":
            final_categories = [
                "NVR",
                "Storage HDD",
                "Server",
                "Windows Server License",
                "Workstation",
                "Monitor",
                "VMS Base License",
                "VMS Channel License",
                "Testing & Commissioning",
            ]
            relationship_type = "Technical Refinement + System Breakdown"

        else:
            final_categories = []
            relationship_type = "Needs Review"

        final_matches: list[dict[str, Any]] = []

        for category in final_categories:
            final_matches.extend(
                final_by_category.get(category, [])
            )

        relationships.append(
            {
                "source_item": source_name,
                "source_description": source["description"],
                "source_quantity": source["quantity"],
                "source_unit": source["unit"],
                "rfq_descriptions": "\n".join(
                    record["description"]
                    for record in rfq_matches
                ),
                "rfq_quantities": ", ".join(
                    str(record["quantity"])
                    for record in rfq_matches
                ),
                "final_categories": "\n".join(
                    record["mapped_category"]
                    for record in final_matches
                ),
                "final_part_numbers": "\n".join(
                    record["part_number"]
                    for record in final_matches
                    if record["part_number"]
                ),
                "final_quantities": ", ".join(
                    str(record["quantity"])
                    for record in final_matches
                    if record["quantity"] is not None
                ),
                "relationship_type": relationship_type,
                "engineering_observation": "",
                "approved_learning_rule": "",
            }
        )

    return relationships


def total_quantity(
    records: list[dict[str, Any]],
    names: set[str],
    field: str,
) -> float:
    total = 0.0

    for record in records:
        if record.get(field) in names:
            quantity = record.get("quantity")

            if quantity is not None:
                total += float(quantity)

    return total


def validate_results(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
    relationships: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    camera_names = {
        "Indoor Dome Camera",
        "Outdoor Dome Camera",
        "Outdoor Bullet Camera",
        "Anti-fog Bullet Camera",
        "Anti-fog Bullet Camera with Pole",
        "Anti-fog Dome Camera",
    }

    final_camera_categories = {
        "Dome Camera Consolidated",
        "Outdoor Bullet Camera",
        "Anti-fog Bullet Camera",
        "Anti-fog Dome Camera",
    }

    source_camera_qty = total_quantity(
        source_records,
        camera_names,
        "source_item",
    )

    rfq_camera_qty = total_quantity(
        rfq_records,
        camera_names,
        "mapped_source_item",
    )

    final_camera_qty = total_quantity(
        final_records,
        final_camera_categories,
        "mapped_category",
    )

    nvr_qty = total_quantity(
        final_records,
        {"NVR"},
        "mapped_category",
    )

    workstation_qty = total_quantity(
        final_records,
        {"Workstation"},
        "mapped_category",
    )

    checks = [
        {
            "check": "Source measurable items",
            "expected": 7,
            "actual": len(source_records),
            "status": "PASS" if len(source_records) == 7 else "FAIL",
        },
        {
            "check": "Source camera quantity",
            "expected": 213,
            "actual": source_camera_qty,
            "status": "PASS" if source_camera_qty == 213 else "FAIL",
        },
        {
            "check": "RFQ camera quantity",
            "expected": 213,
            "actual": rfq_camera_qty,
            "status": "PASS" if rfq_camera_qty == 213 else "FAIL",
        },
        {
            "check": "Final camera quantity",
            "expected": 213,
            "actual": final_camera_qty,
            "status": "PASS" if final_camera_qty == 213 else "FAIL",
        },
        {
            "check": "Relationship rows",
            "expected": 7,
            "actual": len(relationships),
            "status": "PASS" if len(relationships) == 7 else "FAIL",
        },
        {
            "check": "Final NVR quantity",
            "expected": 1,
            "actual": nvr_qty,
            "status": "PASS" if nvr_qty == 1 else "FAIL",
        },
        {
            "check": "Final Workstation quantity",
            "expected": 1,
            "actual": workstation_qty,
            "status": "PASS" if workstation_qty == 1 else "FAIL",
        },
    ]

    return checks


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append(
            [
                record.get(header, "")
                for header in headers
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            value = clean(cell.value)
            max_length = max(max_length, len(value))

        width = min(max(max_length + 2, 12), 55)
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = width


def main() -> None:
    for path in (SOURCE_BOQ, ENGINEER_RFQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_cctv()
    rfq_records = extract_engineer_rfq()
    final_records = extract_final_cctv()

    relationships = build_relationships(
        source_records,
        rfq_records,
        final_records,
    )

    checks = validate_results(
        source_records,
        rfq_records,
        final_records,
        relationships,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Source CCTV BOQ"),
        [
            "source_item",
            "sheet",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        source_records,
    )

    write_sheet(
        workbook.create_sheet("Engineer CCTV RFQ"),
        [
            "mapped_source_item",
            "sheet",
            "source_row",
            "part_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        rfq_records,
    )

    write_sheet(
        workbook.create_sheet("Final Quotation CCTV"),
        [
            "page",
            "mapped_category",
            "part_number",
            "description",
            "quantity",
            "unit_price",
            "item_price",
        ],
        final_records,
    )

    write_sheet(
        workbook.create_sheet("Relationships"),
        [
            "source_item",
            "source_description",
            "source_quantity",
            "source_unit",
            "rfq_descriptions",
            "rfq_quantities",
            "final_categories",
            "final_part_numbers",
            "final_quantities",
            "relationship_type",
            "engineering_observation",
            "approved_learning_rule",
        ],
        relationships,
    )

    write_sheet(
        workbook.create_sheet("Validation"),
        [
            "check",
            "expected",
            "actual",
            "status",
        ],
        checks,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source CCTV measurable items: {len(source_records)}")
    print(f"Engineer RFQ measurable rows: {len(rfq_records)}")
    print(f"Final CCTV line items: {len(final_records)}")
    print(f"Relationship rows: {len(relationships)}")
    print()

    for check in checks:
        print(
            f'{check["status"]}: {check["check"]} '
            f'expected={check["expected"]} actual={check["actual"]}'
        )

    print()
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
