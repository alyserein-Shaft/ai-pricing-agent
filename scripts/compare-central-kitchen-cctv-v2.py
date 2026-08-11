from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path("inputs/central-kitchen")

SOURCE_BOQ = BASE / "source-boq.xlsx"
ENGINEER_RFQ = BASE / "cctv-engineer-rfq.xlsx"
FINAL_QUOTATION = (
    BASE
    / "final-quotation"
    / "Q1067-626-LCU- Central Kitchen - Makkah.pdf"
)

OUTPUT_DIR = Path("outputs/central-kitchen-cctv-comparison")
OUTPUT_FILE = OUTPUT_DIR / "Central_Kitchen_CCTV_Comparison_v2.xlsx"


SOURCE_ITEMS = [
    {
        "source_item": "NVR",
        "keywords": ["network video recorder", "nvr"],
    },
    {
        "source_item": "Indoor Dome Camera",
        "keywords": ["indoor dome"],
    },
    {
        "source_item": "Outdoor Dome Camera",
        "keywords": ["outdoor dome"],
    },
    {
        "source_item": "Outdoor Bullet Camera",
        "keywords": ["outdoor bullet"],
    },
    {
        "source_item": "Anti-fog Bullet Camera",
        "keywords": ["anti fog bullet", "antifog bullet"],
    },
    {
        "source_item": "Anti-fog Bullet Camera with Pole",
        "keywords": [
            "anti fog bullet camera with pole",
            "antifog bullet camera with pole",
            "bullet camera with pole",
        ],
    },
    {
        "source_item": "Anti-fog Dome Camera",
        "keywords": ["anti fog dome", "antifog dome"],
    },
]


def clean(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()
    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean(value).replace(",", "")

    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)

    return None


def row_values(ws, row_number: int) -> list[Any]:
    return [
        ws.cell(row=row_number, column=column).value
        for column in range(1, ws.max_column + 1)
    ]


def longest_text(values: list[Any]) -> str:
    candidates = [
        clean(value)
        for value in values
        if isinstance(value, str) and len(clean(value)) >= 8
    ]

    if not candidates:
        return ""

    return max(candidates, key=len)


def find_quantity(values: list[Any]) -> float | None:
    candidates = []

    for index, value in enumerate(values):
        number = numeric(value)

        if number is not None:
            candidates.append((index, number))

    if not candidates:
        return None

    for _, number in reversed(candidates):
        if number >= 1:
            return number

    return candidates[-1][1]


def find_unit(values: list[Any]) -> str:
    known_units = {
        "no",
        "nos",
        "no.",
        "each",
        "ea",
        "set",
        "lot",
        "ls",
        "point",
        "points",
    }

    for value in values:
        text = normalize(value)

        if text in known_units:
            return clean(value)

    return ""


def find_item_number(values: list[Any]) -> str:
    for value in values[:6]:
        text = clean(value)

        if re.fullmatch(r"\d{1,3}(?:\.\d{1,3})+", text):
            return text

    return ""


def match_source_item(description: str) -> str:
    normalized = normalize(description)

    for item in SOURCE_ITEMS:
        for keyword in item["keywords"]:
            if normalize(keyword) in normalized:
                return item["source_item"]

    return ""


def extract_source_cctv() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_BOQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if "28" not in sheet_name and "cctv" not in sheet_name.lower():
            continue

        ws = workbook[sheet_name]
        inside_cctv = False

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))
            normalized = normalize(joined)

            if not joined:
                continue

            if any(
                phrase in normalized
                for phrase in (
                    "closed circuit television",
                    "cctv system",
                    "video surveillance",
                )
            ):
                inside_cctv = True

            if inside_cctv and any(
                phrase in normalized
                for phrase in (
                    "card access system",
                    "access control system",
                    "fire alarm system",
                    "public address system",
                    "data network",
                )
            ):
                if "closed circuit television" not in normalized:
                    break

            if not inside_cctv:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)
            unit = find_unit(values)
            item_number = find_item_number(values)
            source_item = match_source_item(description)

            if not source_item:
                continue

            if quantity is None:
                continue

            records.append(
                {
                    "source_item": source_item,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "item_number": item_number,
                    "description": description,
                    "quantity": quantity,
                    "unit": unit,
                    "raw_row": joined,
                }
            )

    return records


def extract_engineer_rfq() -> list[dict[str, Any]]:
    workbook = load_workbook(ENGINEER_RFQ, data_only=False)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for row_number in range(1, ws.max_row + 1):
            values = row_values(ws, row_number)
            joined = clean(" ".join(clean(value) for value in values))

            if not joined:
                continue

            description = longest_text(values)
            quantity = find_quantity(values)
            unit = find_unit(values)

            if not description or quantity is None:
                continue

            part_number = ""

            for value in values:
                text = clean(value)

                if (
                    len(text) >= 4
                    and re.search(r"[A-Za-z]", text)
                    and re.search(r"\d", text)
                    and " " not in text
                    and not text.startswith("=")
                ):
                    part_number = text
                    break

            source_item = match_source_item(description)

            if not source_item:
                normalized = normalize(description)

                if "workstation" in normalized:
                    source_item = "Workstation"
                elif "monitor" in normalized:
                    source_item = "Monitor"
                elif "server" in normalized:
                    source_item = "Server"
                else:
                    source_item = "Engineer Added"

            records.append(
                {
                    "mapped_source_item": source_item,
                    "sheet": sheet_name,
                    "source_row": row_number,
                    "part_number": part_number,
                    "description": description,
                    "quantity": quantity,
                    "unit": unit,
                    "raw_row": joined,
                }
            )

    return records


def is_item_start(line: str) -> bool:
    return bool(
        re.match(
            r"^\s*\d+\s+",
            line,
        )
    )


def parse_final_item(block: list[str], page_number: int) -> dict[str, Any]:
    full_text = clean(" ".join(block))

    serial_number = ""
    serial_match = re.match(r"^\s*(\d+)\s+", full_text)

    if serial_match:
        serial_number = serial_match.group(1)

    part_number = ""

    part_matches = re.findall(
        r"\b[A-Z0-9]+(?:[-/][A-Z0-9()]+)+\b",
        full_text,
        flags=re.IGNORECASE,
    )

    if part_matches:
        part_number = part_matches[0]

    quantity = None
    unit_price = None
    item_price = None

    amount_match = re.search(
        r"\b(\d+(?:\.\d+)?)\s+"
        r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s+"
        r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*$",
        full_text,
    )

    if amount_match:
        quantity = float(amount_match.group(1))
        unit_price = float(amount_match.group(2).replace(",", ""))
        item_price = float(amount_match.group(3).replace(",", ""))

    description = full_text

    if serial_number:
        description = re.sub(
            rf"^\s*{re.escape(serial_number)}\s+",
            "",
            description,
            count=1,
        )

    mapped_category = map_final_category(description)

    return {
        "serial_number": serial_number,
        "page": page_number,
        "part_number": part_number,
        "description": description,
        "quantity": quantity,
        "unit_price": unit_price,
        "item_price": item_price,
        "mapped_category": mapped_category,
    }


def map_final_category(description: str) -> str:
    normalized = normalize(description)

    if "nvr" in normalized or "network video recorder" in normalized:
        return "NVR"

    if "hard disk" in normalized or "hdd" in normalized:
        return "Storage HDD"

    if "junction box" in normalized:
        return "Junction Box"

    if "pole mount" in normalized or "pole mounting" in normalized:
        return "Pole Mount"

    if "defog" in normalized and "bullet" in normalized:
        return "Anti-fog Bullet Camera"

    if "defog" in normalized and "dome" in normalized:
        return "Anti-fog Dome Camera"

    if "bullet camera" in normalized or "bullet network camera" in normalized:
        return "Outdoor Bullet Camera"

    if "dome camera" in normalized or "dome network camera" in normalized:
        return "Dome Camera Consolidated"

    if "workstation" in normalized:
        return "Workstation"

    if "monitor" in normalized:
        return "Monitor"

    if "windows server" in normalized:
        return "Windows Server License"

    if "server" in normalized:
        return "Server"

    if "license" in normalized and "channel" in normalized:
        return "VMS Channel License"

    if "license" in normalized:
        return "VMS Base License"

    if (
        "testing" in normalized
        or "commissioning" in normalized
        or "configuration" in normalized
    ):
        return "Testing & Commissioning"

    return "Other CCTV Component"


def extract_final_cctv() -> list[dict[str, Any]]:
    reader = PdfReader(FINAL_QUOTATION)
    records: list[dict[str, Any]] = []

    inside_cctv = False
    current_block: list[str] = []
    current_page = 0

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""

        for raw_line in text.splitlines():
            line = clean(raw_line)

            if not line:
                continue

            normalized = normalize(line)

            if any(
                phrase in normalized
                for phrase in (
                    "closed circuit television",
                    "cctv system",
                    "cctv",
                )
            ):
                inside_cctv = True

            if inside_cctv and any(
                phrase in normalized
                for phrase in (
                    "access control system",
                    "public address system",
                    "fire alarm system",
                    "data networking",
                    "structured cabling",
                )
            ):
                if "cctv" not in normalized:
                    if current_block:
                        records.append(
                            parse_final_item(current_block, current_page)
                        )
                        current_block = []

                    inside_cctv = False
                    continue

            if not inside_cctv:
                continue

            if is_item_start(line):
                if current_block:
                    records.append(
                        parse_final_item(current_block, current_page)
                    )

                current_block = [line]
                current_page = page_number
            elif current_block:
                current_block.append(line)

    if current_block:
        records.append(parse_final_item(current_block, current_page))

    return [
        record
        for record in records
        if record["serial_number"]
        and record["mapped_category"] != "Other CCTV Component"
    ]


def build_relationships(
    source_records: list[dict[str, Any]],
    rfq_records: list[dict[str, Any]],
    final_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    relationships: list[dict[str, Any]] = []

    source_by_name = {
        record["source_item"]: record
        for record in source_records
    }

    rfq_by_name: dict[str, list[dict[str, Any]]] = {}

    for record in rfq_records:
        rfq_by_name.setdefault(
            record["mapped_source_item"],
            [],
        ).append(record)

    final_by_category: dict[str, list[dict[str, Any]]] = {}

    for record in final_records:
        final_by_category.setdefault(
            record["mapped_category"],
            [],
        ).append(record)

    for source_name, source in source_by_name.items():
        rfq_matches = rfq_by_name.get(source_name, [])

        final_categories = []

        if source_name in {"Indoor Dome Camera", "Outdoor Dome Camera"}:
            final_categories = ["Dome Camera Consolidated", "Junction Box"]

        elif source_name == "Outdoor Bullet Camera":
            final_categories = ["Outdoor Bullet Camera", "Junction Box"]

        elif source_name == "Anti-fog Bullet Camera":
            final_categories = ["Anti-fog Bullet Camera"]

        elif source_name == "Anti-fog Bullet Camera with Pole":
            final_categories = [
                "Anti-fog Bullet Camera",
                "Pole Mount",
            ]

        elif source_name == "Anti-fog Dome Camera":
            final_categories = [
                "Anti-fog Dome Camera",
                "Junction Box",
            ]

        elif source_name == "NVR":
            final_categories = [
                "NVR",
                "Storage HDD",
                "Server",
                "Windows Server License",
                "Workstation",
                "Monitor",
                "VMS Base License",
                "VMS Channel License",
                "Testing & Commissioning",
            ]

        final_matches = []

        for category in final_categories:
            final_matches.extend(final_by_category.get(category, []))

        if source_name in {"Indoor Dome Camera", "Outdoor Dome Camera"}:
            relation_type = "Technical Consolidation"

        elif source_name == "NVR":
            relation_type = "Technical Refinement + System Breakdown"

        elif len(final_matches) > 1:
            relation_type = "Product + Accessories Breakdown"

        else:
            relation_type = "Direct Product Mapping"

        relationships.append(
            {
                "source_item": source_name,
                "source_description": source["description"],
                "source_quantity": source["quantity"],
                "source_unit": source["unit"],
                "rfq_descriptions": "\n".join(
                    record["description"]
                    for record in rfq_matches
                ),
                "rfq_quantities": ", ".join(
                    str(record["quantity"])
                    for record in rfq_matches
                ),
                "final_categories": "\n".join(
                    record["mapped_category"]
                    for record in final_matches
                ),
                "final_part_numbers": "\n".join(
                    record["part_number"]
                    for record in final_matches
                    if record["part_number"]
                ),
                "final_quantities": ", ".join(
                    str(record["quantity"])
                    for record in final_matches
                    if record["quantity"] is not None
                ),
                "relationship_type": relation_type,
                "engineering_observation": "",
                "approved_learning_rule": "",
            }
        )

    return relationships


def write_sheet(
    ws,
    headers: list[str],
    records: list[dict[str, Any]],
) -> None:
    ws.append(headers)

    for record in records:
        ws.append(
            [
                record.get(header, "")
                for header in headers
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            value = clean(cell.value)
            max_length = max(max_length, len(value))

        width = min(max(max_length + 2, 12), 55)
        letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[letter].width = width


def main() -> None:
    for path in (SOURCE_BOQ, ENGINEER_RFQ, FINAL_QUOTATION):
        if not path.exists():
            raise FileNotFoundError(f"Missing input: {path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_records = extract_source_cctv()
    rfq_records = extract_engineer_rfq()
    final_records = extract_final_cctv()

    relationships = build_relationships(
        source_records,
        rfq_records,
        final_records,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook.create_sheet("Source CCTV BOQ"),
        [
            "source_item",
            "sheet",
            "source_row",
            "item_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        source_records,
    )

    write_sheet(
        workbook.create_sheet("Engineer CCTV RFQ"),
        [
            "mapped_source_item",
            "sheet",
            "source_row",
            "part_number",
            "description",
            "quantity",
            "unit",
            "raw_row",
        ],
        rfq_records,
    )

    write_sheet(
        workbook.create_sheet("Final Quotation CCTV"),
        [
            "serial_number",
            "page",
            "mapped_category",
            "part_number",
            "description",
            "quantity",
            "unit_price",
            "item_price",
        ],
        final_records,
    )

    write_sheet(
        workbook.create_sheet("Relationships"),
        [
            "source_item",
            "source_description",
            "source_quantity",
            "source_unit",
            "rfq_descriptions",
            "rfq_quantities",
            "final_categories",
            "final_part_numbers",
            "final_quantities",
            "relationship_type",
            "engineering_observation",
            "approved_learning_rule",
        ],
        relationships,
    )

    workbook.save(OUTPUT_FILE)

    print(f"Source CCTV measurable items: {len(source_records)}")
    print(f"Engineer RFQ measurable rows: {len(rfq_records)}")
    print(f"Final CCTV line items: {len(final_records)}")
    print(f"Relationship rows: {len(relationships)}")
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
