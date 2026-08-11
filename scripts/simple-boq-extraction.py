from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_ALIASES = {
    "item": {"item", "item no", "item number", "s no", "s/n"},
    "description": {"description", "item description", "scope description"},
    "quantity": {"quantity", "qty", "boq qty"},
    "unit": {"unit", "uom", "unit of measure"},
}

EXCEL_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


@dataclass
class ColumnMap:
    header_row: int
    item: int | None
    description: int | None
    quantity: int | None
    unit: int | None


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def find_header(ws) -> ColumnMap | None:
    best: tuple[int, ColumnMap] | None = None

    for row_number in range(1, min(ws.max_row, 50) + 1):
        found: dict[str, int] = {}

        for column in range(1, min(ws.max_column, 30) + 1):
            text = normalize(ws.cell(row_number, column).value)
            if not text:
                continue

            for field, aliases in HEADER_ALIASES.items():
                if text in aliases:
                    found[field] = column

        mapping = ColumnMap(
            header_row=row_number,
            item=found.get("item"),
            description=found.get("description"),
            quantity=found.get("quantity"),
            unit=found.get("unit"),
        )

        score = len(found)
        if best is None or score > best[0]:
            best = (score, mapping)

    if not best or best[0] < 3:
        return None

    return best[1]


def is_numeric(value: Any) -> bool:
    if isinstance(value, bool):
        return False

    if isinstance(value, (int, float)):
        return True

    if isinstance(value, str):
        try:
            float(value.replace(",", "").strip())
            return True
        except ValueError:
            return False

    return False


def format_item_number(value: Any) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, str):
        if value in EXCEL_ERRORS:
            return ""
        return value.strip()

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        rounded = round(value, 6)
        if rounded.is_integer():
            return str(int(rounded))
        return f"{rounded:.6f}".rstrip("0").rstrip(".")

    return str(value)


def classify_row(
    values: list[Any],
    item: Any,
    description: Any,
    quantity: Any,
    unit: Any,
) -> tuple[str, str]:
    populated = [
        value for value in values
        if value not in (None, "") and value not in EXCEL_ERRORS
    ]

    if not populated:
        if any(value in EXCEL_ERRORS for value in values):
            return "Formula / Formatting Error", "Excel error without BOQ content"
        return "Blank / Formatting", "No populated cells"

    item_text = normalize(item)
    description_text = normalize(description)
    full_text = normalize(" ".join(str(value) for value in populated))

    if any(term in full_text for term in ("grand total", "total amount")):
        return "Total", "Explicit total wording"

    if "subtotal" in full_text or "carried to collection" in full_text:
        return "Subtotal", "Explicit subtotal or carry-forward wording"

    if item_text.startswith("division ") or description_text.startswith("division "):
        return "Division Header", "Division heading detected"

    if item_text.startswith("section ") or description_text.startswith("section "):
        return "Section", "Section heading detected"

    if full_text == "collection" or description_text == "collection":
        return "Collection / Carry Forward", "Collection row"

    if full_text.startswith("from page") or description_text.startswith("from page"):
        return "Reference Row", "Cross-page reference row"

    if description_text in {"description", "electrical works"}:
        return "Header", "Workbook heading"

    if description and is_numeric(quantity) and unit not in (None, ""):
        return "BOQ Item", "Description, numeric quantity and unit are present"

    if description and quantity in (None, ""):
        text_length = len(str(description).strip())

        scope_terms = (
            "including the supply",
            "supply and install of",
            "complete system",
            "as specified",
            "consultant engineer instructions",
        )

        if text_length > 160 or any(term in description_text for term in scope_terms):
            return "Scope / Preamble", "Long non-measurable scope statement"

        return "Section", "Description without measurable quantity"

    return "Needs Review", "Row does not satisfy deterministic rules"


def update_hierarchy(
    row_type: str,
    item: Any,
    description: Any,
    hierarchy: dict[str, str],
) -> None:
    text = str(description or item or "").strip()
    if not text:
        return

    if row_type == "Division Header":
        hierarchy["division"] = text
        hierarchy["section"] = ""
        hierarchy["system"] = ""
        return

    if row_type == "Section":
        normalized = normalize(text)

        if normalized.startswith("section "):
            hierarchy["section"] = text
            hierarchy["system"] = ""
        elif "system" in normalized or "boxes" in normalized:
            hierarchy["system"] = text
        elif not hierarchy["section"]:
            hierarchy["section"] = text


def hierarchy_text(hierarchy: dict[str, str]) -> str:
    return " > ".join(
        value for value in (
            hierarchy["division"],
            hierarchy["section"],
            hierarchy["system"],
        )
        if value
    )


def extract_boq(input_path: Path) -> list[dict[str, Any]]:
    workbook_formulas = load_workbook(input_path, data_only=False, read_only=True)
    workbook_values = load_workbook(input_path, data_only=True, read_only=True)

    output: list[dict[str, Any]] = []

    for ws_formula in workbook_formulas.worksheets:
        ws_values = workbook_values[ws_formula.title]
        mapping = find_header(ws_formula)

        if mapping is None:
            continue

        hierarchy = {
            "division": "",
            "section": "",
            "system": "",
        }

        for row_number in range(1, ws_formula.max_row + 1):
            values = [
                ws_formula.cell(row_number, column).value
                for column in range(1, ws_formula.max_column + 1)
            ]

            item_formula_raw = (
                ws_formula.cell(row_number, mapping.item).value
                if mapping.item else None
            )

            item_value_raw = (
                ws_values.cell(row_number, mapping.item).value
                if mapping.item else None
            )

            description = (
                ws_formula.cell(row_number, mapping.description).value
                if mapping.description else None
            )

            quantity = (
                ws_values.cell(row_number, mapping.quantity).value
                if mapping.quantity else None
            )

            unit = (
                ws_formula.cell(row_number, mapping.unit).value
                if mapping.unit else None
            )

            item_formula = ""
            formula_status = ""

            if isinstance(item_formula_raw, str):
                if item_formula_raw.startswith("="):
                    item_formula = item_formula_raw
                elif item_formula_raw in EXCEL_ERRORS:
                    formula_status = item_formula_raw

            if isinstance(item_value_raw, str) and item_value_raw in EXCEL_ERRORS:
                formula_status = item_value_raw
                item_value_raw = None

            row_type, reason = classify_row(
                values,
                item_value_raw or item_formula_raw,
                description,
                quantity,
                unit,
            )

            update_hierarchy(
                row_type,
                item_value_raw or item_formula_raw,
                description,
                hierarchy,
            )

            output.append({
                "project": "Central Kitchen Project — Makkah",
                "source_file": input_path.name,
                "sheet": ws_formula.title,
                "source_row": row_number,
                "item_number": format_item_number(item_value_raw),
                "item_formula": item_formula,
                "formula_status": formula_status,
                "description": description,
                "quantity": quantity,
                "unit": unit,
                "hierarchy": hierarchy_text(hierarchy),
                "row_type": row_type,
                "classification_reason": reason,
                "needs_review": "Yes" if row_type == "Needs Review" else "No",
            })

    return output


def write_output(rows: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    headers = [
        "Project",
        "Source File",
        "Sheet",
        "Source Row",
        "Item Number",
        "Item Formula",
        "Formula Status",
        "Description",
        "Quantity",
        "Unit",
        "Hierarchy",
        "Row Type",
        "Classification Reason",
        "Needs Review",
        "Reviewer Decision",
        "Reviewer Notes",
    ]

    def add_sheet(name: str, selected_rows: list[dict[str, Any]]) -> None:
        ws = workbook.create_sheet(name)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(selected_rows) + 1, 1)}"

        for column, header in enumerate(headers, 1):
            cell = ws.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B2A4A")
            cell.alignment = Alignment(wrap_text=True)

        for row_number, row in enumerate(selected_rows, 2):
            values = [
                row["project"],
                row["source_file"],
                row["sheet"],
                row["source_row"],
                row["item_number"],
                row["item_formula"],
                row["formula_status"],
                row["description"],
                row["quantity"],
                row["unit"],
                row["hierarchy"],
                row["row_type"],
                row["classification_reason"],
                row["needs_review"],
                "",
                "",
            ]

            for column, value in enumerate(values, 1):
                cell = ws.cell(row_number, column, value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        widths = {
            1: 28,
            2: 34,
            3: 14,
            4: 12,
            5: 14,
            6: 20,
            7: 16,
            8: 70,
            9: 14,
            10: 12,
            11: 55,
            12: 24,
            13: 38,
            14: 14,
            15: 20,
            16: 40,
        }

        for column, width in widths.items():
            ws.column_dimensions[get_column_letter(column)].width = width

    clean_rows = [
        row for row in rows
        if row["row_type"] == "BOQ Item"
    ]

    review_rows = [
        row for row in rows
        if row["row_type"] == "Needs Review"
    ]

    add_sheet("Clean BOQ", clean_rows)
    add_sheet("Needs Review", review_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "outputs/simple-boq-extraction/Central_Kitchen_BOQ_Review_v3.xlsx"
        ),
    )
    arguments = parser.parse_args()

    if not arguments.input.exists():
        raise SystemExit(f"Input file not found: {arguments.input}")

    rows = extract_boq(arguments.input)
    write_output(rows, arguments.output)

    counts: dict[str, int] = {}
    for row in rows:
        counts[row["row_type"]] = counts.get(row["row_type"], 0) + 1

    print(json.dumps({
        "input": str(arguments.input),
        "output": str(arguments.output),
        "total_rows": len(rows),
        "row_types": counts,
        "clean_boq_items": sum(row["row_type"] == "BOQ Item" for row in rows),
        "needs_review": sum(row["row_type"] == "Needs Review" for row in rows),
        "sheets": sorted({row["sheet"] for row in rows}),
    }, indent=2))


if __name__ == "__main__":
    main()
